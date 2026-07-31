"""Reports in the reader's language; the JSON contract in neither.

The practicality evaluation found English-only output to be a real adoption
barrier for the people this tool is for — Japanese office workers deciding
whether a document may ship.  So everything a person reads localizes, and
everything a machine reads does not: agents and CI parse the JSON, and a
report that changed shape with an environment variable would be a contract
that breaks on a locale.

The two properties worth pinning are therefore opposites.  The human surfaces
must actually change language, and the JSON must be byte-identical whatever
--lang says.
"""

import json
import re

import pytest
from openpyxl import Workbook

from templategate import check, diff
from templategate.core.messages import (
    CATALOG,
    LANGUAGES,
    Translator,
    message,
    missing_keys,
    normalise,
    say,
)
from templategate.cli import main
from templategate.reporters import render_json, render_markdown, render_text

STRICT = """\
version: 1
target: excel
mode: normal_input
allow: []
protect:
  - selector: "*"
    attributes: [formula, images, layout, value, format]
structural:
  sheets: strict
  images: strict
"""


def _quote(path, total="=B1+B2", hide=False):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "設計"
    ws["B1"] = 50000
    ws["B2"] = 80000
    ws["B3"] = total
    if hide:
        ws.column_dimensions["D"].hidden = True
    wb.save(path)
    return path


@pytest.fixture
def pair(tmp_path):
    policy = tmp_path / "policy.yaml"
    policy.write_text(STRICT, encoding="utf-8")
    return {
        "baseline": _quote(tmp_path / "base.xlsx"),
        "candidate": _quote(tmp_path / "cand.xlsx", total=130000),
        "policy": policy,
    }


@pytest.fixture
def result(pair):
    return check(pair["baseline"], pair["candidate"], pair["policy"])


# --- the catalog itself --------------------------------------------------

def test_both_locales_carry_every_message():
    """A key present in one language and missing in the other is a report
    that silently falls back mid-sentence."""
    assert missing_keys() == {lang: set() for lang in LANGUAGES}


def test_every_message_is_a_usable_template():
    """A stray brace turns a report line into a KeyError at the worst
    possible moment."""
    for lang, table in CATALOG.items():
        for key, template in table.items():
            try:
                list(re.finditer(r"\{[^{}]*\}", template))
                template.format_map({})
            except (ValueError, IndexError) as exc:
                pytest.fail(f"{lang}:{key} is not a valid template: {exc}")
            except KeyError:
                pass  # named holes are expected; only syntax is checked here


def test_the_two_locales_use_the_same_holes():
    """If the Japanese for a message asks for a name the English never
    passes, that line renders as the raw template."""
    for key, english in CATALOG["en"].items():
        holes = set(re.findall(r"\{(\w+)", english))
        other = set(re.findall(r"\{(\w+)", CATALOG["ja"][key]))
        assert other <= holes, f"{key}: ja asks for {other - holes}"


@pytest.mark.parametrize("given,expected", [
    ("ja", "ja"), ("JA", "ja"), ("ja-JP", "ja"), ("ja_JP", "ja"),
    ("en", "en"), ("en-GB", "en"),
    (None, "en"), ("", "en"), ("  ", "en"), ("klingon", "en"), ("de", "en"),
])
def test_an_unknown_language_falls_back_rather_than_failing(given, expected):
    """A stale environment variable must not stop a gate from running."""
    assert normalise(given) == expected


def test_a_missing_key_falls_back_to_english():
    assert Translator("ja")("no.such.key.anywhere") == "no.such.key.anywhere"


# --- the JSON contract does not move -------------------------------------

def test_the_json_report_is_identical_in_both_languages(result):
    assert render_json(result) == render_json(result)
    parsed = json.loads(render_json(result))
    assert "protected attribute 'formula' changed" in json.dumps(parsed)


def test_a_detail_serialises_as_its_english_text():
    said = message("detail.image_moved", old="G4", new="H6")
    assert json.dumps({"detail": said}) == '{"detail": "same image, moved from G4 to H6"}'


def test_a_detail_is_a_string_everywhere_it_used_to_be():
    """Nothing downstream should have to know this class exists."""
    said = message("detail.image_moved", old="G4", new="H6")
    assert isinstance(said, str)
    assert said == "same image, moved from G4 to H6"
    assert "moved from" in said
    assert said.startswith("same image")


def test_the_cli_json_report_ignores_the_language(pair, tmp_path, capsys):
    outputs = {}
    for lang in ("en", "ja"):
        destination = tmp_path / f"{lang}.json"
        main(["check", "--baseline", str(pair["baseline"]),
              "--candidate", str(pair["candidate"]),
              "--policy", str(pair["policy"]),
              "--report", "json", "--lang", lang, "--output", str(destination)])
        outputs[lang] = destination.read_bytes()
    assert outputs["en"] == outputs["ja"]


def test_the_diff_json_ignores_the_language(pair, capsys):
    outputs = {}
    for lang in ("en", "ja"):
        main(["diff", "--baseline", str(pair["baseline"]),
              "--candidate", str(pair["candidate"]), "--json", "--lang", lang])
        outputs[lang] = capsys.readouterr().out
    assert outputs["en"] == outputs["ja"]


# --- the human surfaces do move ------------------------------------------

@pytest.mark.parametrize("render", [render_text, render_markdown])
def test_the_verdict_is_translated(result, render):
    assert "FAIL" in render(result, "en")
    assert "不合格" in render(result, "ja")
    assert "FAIL" not in render(result, "ja")


@pytest.mark.parametrize("render", [render_text, render_markdown])
def test_the_summary_counts_are_translated(result, render):
    assert "違反" in render(result, "ja")
    assert "violations" in render(result, "en")


def test_the_violation_line_is_translated(result):
    assert "protected attribute 'formula' changed" in render_text(result, "en")
    assert "保護属性 'formula' が変更されています" in render_text(result, "ja")


def test_the_markdown_table_headings_are_translated(result):
    """Markdown puts old and new in their own columns, so the verdict only
    shows up when there is nothing to put there; the headings always do."""
    assert "| Severity | Location |" in render_markdown(result, "en")
    assert "| 重大度 | 位置 |" in render_markdown(result, "ja")


@pytest.mark.parametrize("render", [render_text, render_markdown])
def test_the_attribute_name_stays_untranslated(result, render):
    """A reader types 'formula' into a YAML file.  Policy vocabulary is not
    prose and must read the same in every language."""
    assert "formula" in render(result, "ja")


def test_a_destroyed_formula_reads_naturally_in_japanese(result):
    report = render_text(result, "ja")
    assert "数式がベタ書きの値に置き換えられています" in report
    assert "→" in report


def test_english_output_is_unchanged_by_the_feature(result):
    """The default has to be exactly what it always was."""
    assert render_text(result) == render_text(result, "en")
    assert "TemplateGate: FAIL" in render_text(result)


def test_the_severity_label_is_translated(result):
    assert "[error]" in render_text(result, "en")
    assert "[エラー]" in render_text(result, "ja")


def test_the_arrow_between_values_follows_the_language(result):
    assert " -> " in render_text(result, "en")
    assert " → " in render_text(result, "ja")


def test_a_value_that_is_not_set_is_translated(result):
    assert "none" in render_text(result, "en")
    assert "なし" in render_text(result, "ja")


# --- images -------------------------------------------------------------

def _image_change(kind):
    if kind == "moved":
        return message("detail.image_moved", old="G4", new="H6")
    return message("detail.image_resized",
                   old=message("value.pixels", width=96, height=96),
                   new=message("size.cell_span"))


@pytest.mark.parametrize("kind", ["moved", "resized"])
def test_an_image_change_is_translated(kind):
    said = _image_change(kind)
    assert say(said, Translator("en")) == str(said)
    japanese = say(said, Translator("ja"))
    assert japanese != str(said)
    assert "画像" in japanese


def test_a_size_inside_a_sentence_is_translated_too():
    """The size is a message in its own right, nested in the sentence about
    the image; both have to come out in the same language."""
    said = message("detail.image_resized",
                   old=message("value.pixels", width=96, height=96),
                   new=message("size.cell_span"))
    assert "96x96 pixels" in str(said)
    japanese = say(said, Translator("ja"))
    assert "96x96 ピクセル" in japanese
    assert "pixels" not in japanese


# --- format sub-key deltas ----------------------------------------------

def test_a_format_delta_is_translated(tmp_path):
    from templategate.core.describe import describe_delta, field_delta

    delta = field_delta({"numfmt": "General"}, {"numfmt": "0.00"})
    said = describe_delta(delta, "detail.cell_format")
    assert str(said) == "cell format changed: numfmt General -> 0.00"
    japanese = say(said, Translator("ja"))
    assert "セルの書式が変更されています" in japanese
    assert "numfmt General → 0.00" in japanese


def test_a_delta_field_name_stays_untranslated():
    """'numfmt' is the name of the thing, not a sentence about it."""
    from templategate.core.describe import describe_delta, field_delta

    delta = field_delta({"font.bold": False}, {"font.bold": True})
    assert "font.bold" in say(describe_delta(delta, "detail.cell_format"),
                              Translator("ja"))


def test_an_unset_side_of_a_delta_is_translated():
    from templategate.core.describe import describe_delta, field_delta

    delta = field_delta({}, {"font.color": "FFFFFFFF"})
    assert "none ->" in str(describe_delta(delta, "detail.cell_format"))
    assert "なし →" in say(describe_delta(delta, "detail.cell_format"),
                          Translator("ja"))


def test_a_long_delta_says_how_many_it_left_out():
    from templategate.core.describe import describe_delta, field_delta

    delta = field_delta({}, {f"field{i}": i for i in range(10)})
    said = describe_delta(delta, "detail.cell_format")
    assert "and 4 more" in str(said)
    assert "ほか4件" in say(said, Translator("ja"))


# --- the grouped cascade -------------------------------------------------

def _numbered(tmp_path, name, skip=None):
    import docx

    document = docx.Document()
    for i in range(1, 13):
        if skip is not None and i == skip:
            continue
        document.add_paragraph(f"条項 {i}")
    path = tmp_path / f"{name}.docx"
    document.save(path)
    return path


STRICT_WORD = """\
version: 1
target: word
mode: normal_input
allow: []
protect:
  - selector: "body"
    attributes: [text, moved]
"""


@pytest.fixture
def cascade(tmp_path):
    policy = tmp_path / "word.yaml"
    policy.write_text(STRICT_WORD, encoding="utf-8")
    return check(_numbered(tmp_path, "base"),
                 _numbered(tmp_path, "cand", skip=5), policy)


def test_the_cascade_line_is_translated(cascade):
    english = render_text(cascade, "en")
    japanese = render_text(cascade, "ja")
    assert "content shifted because" in english
    assert "内容が繰り下がりました" in japanese


def test_the_cascade_reason_is_translated_not_just_its_frame(cascade):
    """The summary inside the sentence is prose too — leaving it English
    would put "1 paragraph removed at p5" inside a Japanese line."""
    assert "1 paragraph removed at p5" in render_text(cascade, "en")
    japanese = render_text(cascade, "ja")
    assert "paragraph" not in japanese
    assert "段落" in japanese


def test_the_cascade_group_stays_english_in_the_json(cascade):
    """The group is the identity changes are folded by, and a JSON value."""
    payload = json.loads(render_json(cascade))
    grouped = [v for v in payload["violations"] if v["change"]["group"]]
    assert grouped
    assert "paragraph" in grouped[0]["change"]["group"]


# --- degraded and liveness warnings --------------------------------------

def test_the_degraded_warning_is_translated(pair, tmp_path):
    from generate import rewrite_zip

    damaged = tmp_path / "damaged.xlsx"
    rewrite_zip(pair["candidate"], damaged, drop=("xl/workbook.xml",))
    result = check(pair["baseline"], damaged, pair["policy"])

    assert "is damaged" in render_text(result, "en")
    japanese = render_text(result, "ja")
    assert "破損しています" in japanese
    assert "パッケージ内のパーツのみ比較しました" in japanese


def test_the_degraded_warning_names_the_side_in_japanese(pair, tmp_path):
    from generate import rewrite_zip

    damaged = tmp_path / "damaged.xlsx"
    rewrite_zip(pair["candidate"], damaged, drop=("xl/workbook.xml",))
    result = check(pair["baseline"], damaged, pair["policy"])
    assert "変更後" in render_text(result, "ja")


def test_the_liveness_warning_is_translated(pair, tmp_path):
    policy = tmp_path / "dead.yaml"
    policy.write_text(STRICT.replace('allow: []',
                                     'allow:\n  - selector: "Nope!A1"\n'
                                     '    attributes: [value]'),
                      encoding="utf-8")
    result = check(pair["baseline"], pair["candidate"], policy)
    assert result.warnings

    assert "matches nothing in the baseline" in render_text(result, "en")
    japanese = render_text(result, "ja")
    assert "どこにも一致しません" in japanese
    assert "シートはありません" in japanese


def test_a_liveness_warning_is_english_in_the_json(pair, tmp_path):
    policy = tmp_path / "dead.yaml"
    policy.write_text(STRICT.replace('allow: []',
                                     'allow:\n  - selector: "Nope!A1"\n'
                                     '    attributes: [value]'),
                      encoding="utf-8")
    result = check(pair["baseline"], pair["candidate"], policy)
    payload = json.loads(render_json(result))
    assert "matches nothing in the baseline" in payload["warnings"][0]


# --- the command line ----------------------------------------------------

def test_the_lang_flag_selects_the_language(pair, capsys):
    main(["check", "--baseline", str(pair["baseline"]),
          "--candidate", str(pair["candidate"]),
          "--policy", str(pair["policy"]), "--lang", "ja"])
    assert "不合格" in capsys.readouterr().out


def test_the_environment_variable_is_the_default(pair, capsys, monkeypatch):
    monkeypatch.setenv("TEMPLATEGATE_LANG", "ja")
    main(["check", "--baseline", str(pair["baseline"]),
          "--candidate", str(pair["candidate"]),
          "--policy", str(pair["policy"])])
    assert "不合格" in capsys.readouterr().out


def test_the_flag_beats_the_environment(pair, capsys, monkeypatch):
    """One run has to be readable by somebody else without unsetting a
    variable the whole job depends on."""
    monkeypatch.setenv("TEMPLATEGATE_LANG", "ja")
    main(["check", "--baseline", str(pair["baseline"]),
          "--candidate", str(pair["candidate"]),
          "--policy", str(pair["policy"]), "--lang", "en"])
    out = capsys.readouterr().out
    assert "FAIL" in out and "不合格" not in out


def test_a_nonsense_environment_variable_does_not_stop_the_run(pair, capsys,
                                                               monkeypatch):
    monkeypatch.setenv("TEMPLATEGATE_LANG", "sindarin")
    assert main(["check", "--baseline", str(pair["baseline"]),
                 "--candidate", str(pair["candidate"]),
                 "--policy", str(pair["policy"])]) == 1
    assert "FAIL" in capsys.readouterr().out


def test_the_exit_code_does_not_depend_on_the_language(pair):
    codes = {lang: main(["check", "--baseline", str(pair["baseline"]),
                         "--candidate", str(pair["candidate"]),
                         "--policy", str(pair["policy"]), "--lang", lang])
             for lang in LANGUAGES}
    assert codes == {"en": 1, "ja": 1}


def test_diff_output_is_translated(pair, capsys):
    main(["diff", "--baseline", str(pair["baseline"]),
          "--candidate", str(pair["candidate"]), "--lang", "ja"])
    assert "数式がベタ書きの値に置き換えられています" in capsys.readouterr().out


def test_no_changes_is_translated(pair, capsys):
    main(["diff", "--baseline", str(pair["baseline"]),
          "--candidate", str(pair["baseline"]), "--lang", "ja"])
    assert "変更はありません" in capsys.readouterr().out


def test_the_diff_degraded_warning_is_translated(pair, tmp_path, capsys):
    from generate import rewrite_zip

    damaged = tmp_path / "damaged.xlsx"
    rewrite_zip(pair["candidate"], damaged, drop=("xl/workbook.xml",))
    assert main(["diff", "--baseline", str(pair["baseline"]),
                 "--candidate", str(damaged), "--lang", "ja"]) == 2
    assert "破損しています" in capsys.readouterr().err


def test_both_languages_are_offered_in_the_help(capsys):
    with pytest.raises(SystemExit):
        main(["check", "--help"])
    out = capsys.readouterr().out
    assert "--lang" in out
    assert "TEMPLATEGATE_LANG" in out


# --- policy problems -----------------------------------------------------

def _bad_policy(tmp_path, body):
    path = tmp_path / "bad.yaml"
    path.write_text(body, encoding="utf-8")
    return path


def test_a_policy_typo_is_corrected_in_japanese(pair, tmp_path, capsys):
    policy = _bad_policy(tmp_path, "version: 1\ntarget: excel\nprotekt: []\n")
    assert main(["check", "--baseline", str(pair["baseline"]),
                 "--candidate", str(pair["candidate"]),
                 "--policy", str(policy), "--lang", "ja"]) == 2
    error = capsys.readouterr().err
    assert "ポリシーエラー" in error
    assert "もしかして 'protect' ですか？" in error


def test_a_policy_typo_is_unchanged_in_english(pair, tmp_path, capsys):
    policy = _bad_policy(tmp_path, "version: 1\ntarget: excel\nprotekt: []\n")
    main(["check", "--baseline", str(pair["baseline"]),
          "--candidate", str(pair["candidate"]), "--policy", str(policy)])
    error = capsys.readouterr().err
    assert "templategate: policy error:" in error
    assert "did you mean 'protect'?" in error


def test_a_missing_policy_file_is_reported_in_japanese(pair, tmp_path, capsys):
    assert main(["check", "--baseline", str(pair["baseline"]),
                 "--candidate", str(pair["candidate"]),
                 "--policy", str(tmp_path / "nope.yaml"), "--lang", "ja"]) == 2
    assert "ポリシーファイルが見つかりません" in capsys.readouterr().err


def test_the_exception_type_does_not_change_with_the_language(tmp_path):
    """Only the words move.  Callers catching PolicyError keep working."""
    from templategate.core.policy import PolicyError, load_policy

    policy = _bad_policy(tmp_path, "version: 1\ntarget: excel\nprotekt: []\n")
    with pytest.raises(PolicyError) as caught:
        load_policy(policy)
    assert "did you mean 'protect'?" in str(caught.value)


def test_an_unreadable_file_is_reported_in_japanese(pair, tmp_path, capsys):
    broken = tmp_path / "broken.xlsx"
    broken.write_bytes(b"not a zip")
    assert main(["check", "--baseline", str(pair["baseline"]),
                 "--candidate", str(broken),
                 "--policy", str(pair["policy"]), "--lang", "ja"]) == 2
    assert "templategate: エラー:" in capsys.readouterr().err

"""Two programs writing the same document must not look like an edit.

A file authored by a library and then opened and saved by Excel comes back
spelled differently in a dozen small ways — opaque colours gain an alpha,
default attributes appear or vanish, the default font is written onto every
cell — while saying exactly the same thing.  Every one of these was a
systematic false positive found by driving real Excel.

Each class is paired with a real change of the same kind, because a
normalization that quietly swallows the attack is worse than the noise it
removed.  The COM run proved the real files; these reproduce each rewrite as
a targeted transform so CI pins them without needing Excel installed.
"""

import zipfile

import pytest
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from templategate import diff, snapshot


def _attributes(changes):
    return {c.attribute for c in changes}


def _rewrite(source, destination, replacements):
    """Rewrite parts of a package, the way another program's save would."""
    from generate import rewrite_zip

    with zipfile.ZipFile(source) as zf:
        patched = {}
        for name, transform in replacements.items():
            patched[name] = transform(zf.read(name).decode("utf-8")).encode("utf-8")
    rewrite_zip(source, destination, add=patched)
    return destination


# --- F1: an opaque colour is an opaque colour ----------------------------

def _coloured(path, argb):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "total"
    ws["A1"].font = Font(color=argb)
    ws.sheet_properties.tabColor = argb
    wb.save(path)
    return path


def test_alpha_spelling_is_not_a_colour_change(tmp_path):
    """openpyxl writes 00RRGGBB where Excel writes FFRRGGBB."""
    baseline = _coloured(tmp_path / "b.xlsx", "000000FF")
    candidate = _coloured(tmp_path / "c.xlsx", "FF0000FF")
    assert diff(baseline, candidate) == []


def test_a_real_colour_change_still_fails(tmp_path):
    baseline = _coloured(tmp_path / "b.xlsx", "000000FF")
    candidate = _coloured(tmp_path / "c.xlsx", "FFFF0000")
    assert _attributes(diff(baseline, candidate)) == {"format", "sheet_settings"}


def test_conditional_format_fill_alpha_is_normalized(tmp_path):
    def build(path, argb):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.conditional_formatting.add("B2:B4", CellIsRule(
            operator="greaterThan", formula=["8"],
            fill=PatternFill(start_color=argb, end_color=argb, fill_type="solid")))
        wb.save(path)
        return path

    assert diff(build(tmp_path / "b.xlsx", "00FFC7CE"),
                build(tmp_path / "c.xlsx", "FFFFC7CE")) == []
    assert diff(build(tmp_path / "b2.xlsx", "00FFC7CE"),
                build(tmp_path / "c2.xlsx", "FFFFFFFF")) != []


# --- F2: an omitted attribute means its default --------------------------

def _validated(path, **kwargs):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    validation = DataValidation(type="whole", formula1="1", formula2="100",
                                **kwargs)
    ws.add_data_validation(validation)
    validation.add("B2:B4")
    wb.save(path)
    return path


def test_an_omitted_default_equals_the_written_default(tmp_path):
    """Excel drops errorStyle="stop" and operator="between"; they still hold."""
    spelled_out = _validated(tmp_path / "b.xlsx", operator="between",
                             errorStyle="stop")
    omitted = _validated(tmp_path / "c.xlsx")
    assert diff(spelled_out, omitted) == []


def test_a_real_downgrade_is_not_an_omitted_default(tmp_path):
    """stop -> information is the attack; it must never be normalized away."""
    strict = _validated(tmp_path / "b.xlsx", operator="between", errorStyle="stop")
    defanged = _validated(tmp_path / "c.xlsx", operator="between",
                          errorStyle="information")
    assert _attributes(diff(strict, defanged)) == {"data_validation"}


def test_table_part_default_attributes_are_filled(tmp_path):
    """openpyxl spells out headerRowCount; Excel spells out the style flags."""
    from openpyxl.worksheet.table import Table

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row, values in enumerate([("Item", "Qty"), ("a", 1), ("b", 2)], start=1):
        ws.cell(row=row, column=1, value=values[0])
        ws.cell(row=row, column=2, value=values[1])
    ws.add_table(Table(displayName="T1", ref="A1:B3"))
    baseline = tmp_path / "b.xlsx"
    wb.save(baseline)

    def as_excel_would(xml: str) -> str:
        # Drop the default Excel omits, add the defaults Excel writes, and
        # decorate with the revision namespaces Excel stamps on.
        xml = xml.replace(' headerRowCount="1"', "")
        xml = xml.replace("<tableStyleInfo ",
                          '<tableStyleInfo showColumnStripes="0" '
                          'showFirstColumn="0" showLastColumn="0" ')
        return xml.replace("<table ", '<table xr:uid="{DEADBEEF}" '
                           'xmlns:xr="http://schemas.microsoft.com/office/'
                           'spreadsheetml/2014/revision" ')

    candidate = _rewrite(baseline, tmp_path / "c.xlsx",
                         {"xl/tables/table1.xml": as_excel_would})
    assert diff(baseline, candidate) == []


def test_a_real_table_change_still_fails(tmp_path):
    from openpyxl.worksheet.table import Table

    def build(path, name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"], ws["B1"] = "Item", "Qty"
        ws["A2"], ws["B2"] = "a", 1
        ws.add_table(Table(displayName=name, ref="A1:B2"))
        wb.save(path)
        return path

    changes = diff(build(tmp_path / "b.xlsx", "T1"),
                   build(tmp_path / "c.xlsx", "Renamed"))
    assert any(c.attribute == "parts" for c in changes)


# --- F3: the default font is not a font ----------------------------------

def _fonted(path, font):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "total"
    if font is not None:
        ws["A1"].font = font
    wb.save(path)
    return path


def test_materializing_the_default_font_is_not_a_change(tmp_path):
    """Excel writes the locale default onto every cell; that is not styling."""
    unstyled = _fonted(tmp_path / "b.xlsx", None)
    materialized = _fonted(tmp_path / "c.xlsx", Font(name="Calibri", size=11))
    assert diff(unstyled, materialized) == []


def test_a_real_font_change_still_fails(tmp_path):
    unstyled = _fonted(tmp_path / "b.xlsx", None)
    restyled = _fonted(tmp_path / "c.xlsx", Font(name="Papyrus", size=11))
    change = [c for c in diff(unstyled, restyled) if c.attribute == "format"][0]
    assert "font.name" in change.detail


def test_the_default_is_read_from_each_workbook(tmp_path):
    """Two locales disagree about the default; each file is judged by its own."""
    baseline = _fonted(tmp_path / "b.xlsx", None)
    assert snapshot(baseline)["sheets"]["Sheet1"]["cells"]["A1"]["format"] is None


# --- F4: shared strings are storage, not content -------------------------

def test_the_shared_string_table_appearing_is_not_a_change(tmp_path):
    """Excel introduces the shared table on save; the values did not move."""
    from generate import rewrite_zip

    baseline = tmp_path / "b.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"], ws["A2"] = 1, 2          # numbers only: no shared table written
    wb.save(baseline)
    with zipfile.ZipFile(baseline) as zf:
        assert "xl/sharedStrings.xml" not in zf.namelist()

    candidate = tmp_path / "c.xlsx"
    rewrite_zip(baseline, candidate, add={
        "xl/sharedStrings.xml":
            b'<?xml version="1.0"?><sst xmlns="http://schemas.openxmlformats.org'
            b'/spreadsheetml/2006/main" count="0" uniqueCount="0"/>'})
    assert diff(baseline, candidate) == []


# --- F5: a theme change is reported, and explained -----------------------

def test_a_theme_change_is_reported_in_words_a_reader_can_use(tmp_path):
    baseline = tmp_path / "b.xlsx"
    wb = Workbook()
    wb.active["A1"] = "x"
    wb.save(baseline)

    candidate = _rewrite(baseline, tmp_path / "c.xlsx", {
        "xl/theme/theme1.xml": lambda xml: xml.replace("Calibri", "Papyrus")})
    change = [c for c in diff(baseline, candidate) if "theme" in c.location][0]
    assert "theme replaced" in change.detail
    assert "may render differently" in change.detail
    assert "saved by Excel or Word" in change.detail


# --- the writer's own feature list is not the document -------------------

def test_the_saving_programs_feature_list_is_ignored(tmp_path):
    """Excel stamps its calculation-engine features onto every save."""
    baseline = tmp_path / "b.xlsx"
    wb = Workbook()
    wb.active["A1"] = "x"
    wb.save(baseline)

    calc_features = (
        '<extLst><ext uri="{B58B0392-4F1F-4190-BB64-5DF3571DCE5F}" '
        'xmlns:xcalcf="http://schemas.microsoft.com/office/spreadsheetml/2018/'
        'calcfeatures"><xcalcf:calcFeatures><xcalcf:feature name="microsoft.com:RD"/>'
        "</xcalcf:calcFeatures></ext></extLst>"
    )
    candidate = _rewrite(baseline, tmp_path / "c.xlsx", {
        "xl/workbook.xml": lambda xml: xml.replace("</workbook>",
                                                   calc_features + "</workbook>")})
    assert diff(baseline, candidate) == []


def test_a_real_extension_block_is_still_reported(tmp_path):
    """The x14 dropdown lives in an extLst too, and must stay visible."""
    baseline = tmp_path / "b.xlsx"
    wb = Workbook()
    wb.active["A1"] = "x"
    wb.save(baseline)

    x14 = ('<extLst><ext uri="{CCE6A557}"><x14:dataValidations '
           'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main">'
           '<x14:dataValidation type="list"/></x14:dataValidations></ext></extLst>')
    candidate = _rewrite(baseline, tmp_path / "c.xlsx", {
        "xl/worksheets/sheet1.xml":
            lambda xml: xml.replace("</worksheet>", x14 + "</worksheet>")})
    assert [c for c in diff(baseline, candidate) if c.location.endswith("#extLst")]


# --- serialisation is not content ----------------------------------------

def test_xml_declaration_and_namespace_churn_is_not_a_change(tmp_path):
    """Excel rewrites parts with its own declaration and namespace order."""
    baseline = tmp_path / "b.xlsx"
    wb = Workbook()
    wb.active["A1"] = "x"
    wb.save(baseline)

    def as_excel_would(xml: str) -> str:
        # Excel stamps its revision namespaces, an mc:Ignorable and a session
        # uid onto every part it rewrites.
        return xml.replace(
            "<a:theme ",
            '<a:theme mc:Ignorable="xr" xmlns:mc="http://schemas.openxmlformats.org'
            '/markup-compatibility/2006" xmlns:xr="http://schemas.microsoft.com'
            '/office/spreadsheetml/2014/revision" xr:uid="{DEADBEEF-0000}" ', 1)

    candidate = _rewrite(baseline, tmp_path / "c.xlsx",
                         {"xl/theme/theme1.xml": as_excel_would})
    with zipfile.ZipFile(candidate) as zf:
        assert b"mc:Ignorable" in zf.read("xl/theme/theme1.xml")
    assert diff(baseline, candidate) == []

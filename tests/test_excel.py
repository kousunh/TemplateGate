from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from templategate import check, diff, snapshot
from templategate.core.model import (
    ATTR_CONDITIONAL_FORMATTING,
    ATTR_DATA_VALIDATION,
    ATTR_FORMULA,
    ATTR_VALUE,
)
from templategate.core.policy import parse_policy


def test_snapshot_contents(fixtures):
    snap = snapshot(fixtures["excel_baseline"])
    assert snap["target"] == "excel"
    assert set(snap["sheets"]) == {"計画表", "メモ"}
    sheet = snap["sheets"]["計画表"]
    assert sheet["cells"]["A1"]["value"] == "架空市 水道整備計画(サンプル)"
    assert sheet["cells"]["D8"]["formula"] == "=SUM(D4:D6)"
    assert "A1:D1" in sheet["merges"]
    assert sheet["conditional_formatting"]
    assert sheet["data_validation"]
    assert sheet["print"]["orientation"] == "landscape"
    assert sheet["header_footer"]


def test_identical_files_have_no_changes(fixtures):
    assert diff(fixtures["excel_baseline"], fixtures["excel_baseline"]) == []


def test_good_edit_passes(fixtures):
    result = check(fixtures["excel_baseline"], fixtures["excel_good"],
                   fixtures["excel_policy"])
    assert result.passed, [v.message for v in result.violations]
    assert {c.location for c in result.allowed} == {"計画表!B4", "計画表!B5"}
    assert result.semantic_mode == "off"


def test_bad_edit_fails_with_expected_violations(fixtures):
    result = check(fixtures["excel_baseline"], fixtures["excel_bad"],
                   fixtures["excel_policy"])
    assert not result.passed

    by_key = {(v.change.location, v.change.attribute): v for v in result.violations}
    # destroyed formula is caught by the protect rule
    assert by_key[("計画表!D8", "formula")].rule == "protected"
    # header format change is default-denied
    assert by_key[("計画表!A3", "format")].rule == "not_allowed"
    # deleted sheet is a structural violation
    assert ("sheet:メモ", "sheet_structure") in by_key
    # the allowed value edit still goes through
    assert any(c.location == "計画表!B4" for c in result.allowed)


def _sheet_with_values(path, title, values):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for coord, value in values.items():
        ws[coord] = value
    wb.save(path)


def test_renamed_sheet_still_compares_cells(tmp_path):
    """A rename must not carry the sheet's contents out of sight."""
    baseline = tmp_path / "base.xlsx"
    candidate = tmp_path / "cand.xlsx"
    _sheet_with_values(baseline, "Plan", {"A1": "Budget", "B1": 1000, "B2": "=B1*2"})
    _sheet_with_values(candidate, "Plan ", {"A1": "HACKED", "B1": 999999, "B2": 0})

    changes = diff(baseline, candidate)
    by_key = {(c.location, c.attribute) for c in changes}
    assert ("Plan!A1", ATTR_VALUE) in by_key
    assert ("Plan!B1", ATTR_VALUE) in by_key
    assert ("Plan!B2", ATTR_FORMULA) in by_key


def test_sheet_rename_with_sheets_ignore_is_not_a_bypass(tmp_path):
    baseline = tmp_path / "base.xlsx"
    candidate = tmp_path / "cand.xlsx"
    _sheet_with_values(baseline, "Plan", {"A1": "Budget", "B1": 1000, "B2": "=B1*2"})
    _sheet_with_values(candidate, "Plan ", {"A1": "HACKED", "B1": 999999, "B2": 0})

    policy = parse_policy({
        "target": "excel",
        "protect": [{"selector": "*", "attributes": ["*"]}],
        "structural": {"sheets": "ignore"},
    })
    result = check(baseline, candidate, policy)
    assert not result.passed
    assert result.violations


def test_pure_rename_is_reported_as_a_rename(tmp_path):
    baseline = tmp_path / "base.xlsx"
    candidate = tmp_path / "cand.xlsx"
    values = {"A1": "Budget", "B1": 1000}
    _sheet_with_values(baseline, "Plan", values)
    _sheet_with_values(candidate, "Plan2", values)

    changes = diff(baseline, candidate)
    details = {c.location: c.detail for c in changes}
    assert details == {
        "sheet:Plan": "sheet renamed to 'Plan2'",
        "sheet:Plan2": "sheet renamed from 'Plan'",
    }


def test_unrelated_add_and_remove_keeps_neutral_wording(tmp_path):
    """Two unrelated sheets still get compared, but are not called a rename."""
    baseline = tmp_path / "base.xlsx"
    candidate = tmp_path / "cand.xlsx"
    _sheet_with_values(baseline, "Plan", {"A1": "Budget", "B1": 1000})
    _sheet_with_values(candidate, "Notes", {"A1": "totally", "B1": "different"})

    changes = diff(baseline, candidate)
    details = {c.location: c.detail for c in changes if c.location.startswith("sheet:")}
    assert details == {
        "sheet:Plan": "sheet removed; contents compared against 'Notes'",
        "sheet:Notes": "sheet added; contents compared against 'Plan'",
    }
    assert any(c.location == "Plan!A1" for c in changes)


def test_genuinely_added_sheet_is_not_paired(tmp_path):
    baseline = tmp_path / "base.xlsx"
    candidate = tmp_path / "cand.xlsx"
    _sheet_with_values(baseline, "Plan", {"A1": "Budget"})
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan"
    ws["A1"] = "Budget"
    wb.create_sheet("Extra")["A1"] = "new"
    wb.save(candidate)

    changes = diff(baseline, candidate)
    assert [(c.location, c.detail) for c in changes] == [
        ("sheet:Extra", "sheet added")
    ]


def _sheet_with_rules(path, cf_formula, dv_formula):
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.conditional_formatting.add(
        "B2:B5 D2:D5",
        CellIsRule(operator="greaterThan", formula=[cf_formula],
                   fill=PatternFill("solid", fgColor="FFC7CE")),
    )
    dv = DataValidation(type="whole", operator="greaterThan", formula1=dv_formula)
    dv.add("B2:B5")
    dv.add("D2:D5")
    ws.add_data_validation(dv)
    wb.save(path)


def test_multi_range_rules_report_every_range(tmp_path):
    """A sqref like "B2:B5 D2:D5" must not collapse to its first range."""
    baseline = tmp_path / "base.xlsx"
    candidate = tmp_path / "cand.xlsx"
    _sheet_with_rules(baseline, "10", "0")
    _sheet_with_rules(candidate, "999", "500")

    locations = {(c.location, c.attribute) for c in diff(baseline, candidate)}
    assert ("S!B2:B5", ATTR_CONDITIONAL_FORMATTING) in locations
    assert ("S!D2:D5", ATTR_CONDITIONAL_FORMATTING) in locations
    assert ("S!B2:B5", ATTR_DATA_VALIDATION) in locations
    assert ("S!D2:D5", ATTR_DATA_VALIDATION) in locations

    # A policy that only opens up column B must not wave the D rule through.
    policy = parse_policy({
        "target": "excel",
        "allow": [{"selector": "S!B1:B100",
                   "attributes": ["conditional_formatting", "data_validation"]}],
    })
    result = check(baseline, candidate, policy)
    assert not result.passed
    assert {v.change.location for v in result.violations} == {"S!D2:D5"}


def test_array_formula_is_protected_as_a_formula(tmp_path):
    """An array formula is not a str, and used to be filed under `value`.

    A policy that allows value edits then let =A2:A4*C2:C4 become =A2:A4*0.
    """
    from openpyxl.worksheet.formula import ArrayFormula

    baseline = tmp_path / "base.xlsx"
    candidate = tmp_path / "cand.xlsx"
    for path, text in ((baseline, "=A2:A4*C2:C4"), (candidate, "=A2:A4*0")):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A2"], ws["C2"] = 10, 50
        ws["B2"] = ArrayFormula("B2:B4", text)
        wb.save(path)

    changes = diff(baseline, candidate)
    assert [(c.location, c.attribute) for c in changes] == [("Sheet1!B2", ATTR_FORMULA)]

    policy = parse_policy({
        "target": "excel",
        "allow": [{"selector": "*", "attributes": ["value"]}],
        "protect": [{"selector": "*", "attributes": ["formula"]}],
    })
    result = check(baseline, candidate, policy)
    assert not result.passed
    assert result.violations[0].rule == "protected"


def test_array_formula_survives_an_untouched_comparison(tmp_path):
    """ArrayFormula defines no __eq__; comparing a file with itself must be quiet."""
    from openpyxl.worksheet.formula import ArrayFormula

    path = tmp_path / "base.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    ws["D1"] = ArrayFormula("D1:D3", "=A1:A3*2")
    wb.save(path)
    assert diff(path, path) == []


def test_dates_are_stored_as_text(tmp_path):
    """No datetime object may reach the snapshot; JSON has to survive it."""
    import datetime
    import json

    path = tmp_path / "base.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = datetime.datetime(2026, 7, 31, 12, 30)
    ws["A2"] = datetime.time(9, 15)
    wb.save(path)

    cells = snapshot(path)["sheets"]["Sheet1"]["cells"]
    assert cells["A1"]["value"] == "2026-07-31T12:30:00"
    assert cells["A2"]["value"] == "09:15:00"
    json.dumps(snapshot(path))  # no default= crutch needed


def _chartsheet_workbook(path, values):
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for row, value in enumerate(values, start=1):
        ws.cell(row=row, column=1, value=value)
    sheet = wb.create_chartsheet("Chart1")
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=1, min_row=1, max_row=len(values)))
    sheet.add_chart(chart)
    wb.save(path)


def test_chartsheet_does_not_crash(tmp_path):
    """A chartsheet has no cell grid; asking it for rows used to raise."""
    baseline = tmp_path / "base.xlsx"
    _chartsheet_workbook(baseline, [5, 3, 8, 2])

    snap = snapshot(baseline)
    assert snap["sheets"]["Chart1"]["kind"] == "chartsheet"
    assert snap["sheets"]["Chart1"]["cells"] == {}
    assert snap["sheets"]["Data"]["kind"] == "worksheet"
    assert diff(baseline, baseline) == []


def test_chartsheet_cell_edits_on_the_data_sheet_still_report(tmp_path):
    baseline = tmp_path / "base.xlsx"
    candidate = tmp_path / "cand.xlsx"
    _chartsheet_workbook(baseline, [5, 3, 8, 2])
    _chartsheet_workbook(candidate, [5, 3, 8, 99])
    assert [(c.location, c.attribute) for c in diff(baseline, candidate)] == [
        ("Data!A4", ATTR_VALUE)
    ]


def test_worksheet_replaced_by_chartsheet_is_reported(tmp_path):
    baseline = tmp_path / "base.xlsx"
    candidate = tmp_path / "cand.xlsx"
    _sheet_with_values(baseline, "Chart1", {"A1": "was a worksheet"})
    _chartsheet_workbook(candidate, [1, 2])
    details = {c.detail for c in diff(baseline, candidate)}
    assert "sheet kind changed" in details


# --- the surface that does not show up in a cell -------------------------

STRICT_EXCEL = parse_policy({
    "target": "excel",
    "protect": [{"selector": "*", "attributes": ["*"]}],
})


def _edited(tmp_path, build, edit):
    baseline = tmp_path / "base.xlsx"
    candidate = tmp_path / "cand.xlsx"
    build(baseline)
    from openpyxl import load_workbook as _load

    workbook = _load(baseline)
    edit(workbook)
    workbook.save(candidate)
    return baseline, candidate


def _two_rows(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row, (item, qty) in enumerate([("Widget", 10), ("Gadget", 5)], start=2):
        ws[f"A{row}"], ws[f"B{row}"] = item, qty
    wb.save(path)


def _attributes(changes):
    out = {}
    for change in changes:
        out.setdefault(change.attribute, set()).add(change.location)
    return out


def test_hidden_rows_and_columns_are_compared(tmp_path):
    """A hidden row still feeds every formula; it just leaves the page."""
    def hide(wb):
        wb["Sheet1"].row_dimensions[3].hidden = True
        wb["Sheet1"].column_dimensions["B"].hidden = True

    baseline, candidate = _edited(tmp_path, _two_rows, hide)
    assert _attributes(diff(baseline, candidate))["layout"] == {
        "Sheet1!3:3", "Sheet1!B:B"
    }
    assert not check(baseline, candidate, STRICT_EXCEL).passed


def test_column_width_change_is_compared(tmp_path):
    def widen(wb):
        wb["Sheet1"].column_dimensions["A"].width = 3

    baseline, candidate = _edited(tmp_path, _two_rows, widen)
    assert _attributes(diff(baseline, candidate))["layout"] == {"Sheet1!A:A"}


def test_layout_is_addressable_by_range_selector(tmp_path):
    """Row and column locations are ordinary A1 refs, so selectors reach them."""
    def hide(wb):
        wb["Sheet1"].column_dimensions["B"].hidden = True

    baseline, candidate = _edited(tmp_path, _two_rows, hide)
    policy = parse_policy({
        "target": "excel",
        "allow": [{"selector": "Sheet1!B:B", "attributes": ["layout"]}],
    })
    assert check(baseline, candidate, policy).passed


def test_sheet_protection_removal_is_compared(tmp_path):
    def build(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "locked"
        ws.protection.sheet = True
        wb.save(path)

    baseline, candidate = _edited(
        tmp_path, build, lambda wb: setattr(wb["Sheet1"].protection, "sheet", False))
    assert _attributes(diff(baseline, candidate))["protection"] == {
        "Sheet1#protection"
    }
    assert not check(baseline, candidate, STRICT_EXCEL).passed


def test_sheet_scoped_defined_name_is_compared(tmp_path):
    from openpyxl.workbook.defined_name import DefinedName

    def build(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A2"] = 1
        ws.defined_names["LocalRange"] = DefinedName(
            "LocalRange", attr_text="Sheet1!$A$2:$A$4")
        wb.save(path)

    def retarget(wb):
        wb["Sheet1"].defined_names["LocalRange"] = DefinedName(
            "LocalRange", attr_text="Sheet1!$A$2:$A$2")

    baseline, candidate = _edited(tmp_path, build, retarget)
    assert _attributes(diff(baseline, candidate))["defined_names"] == {
        "name:Sheet1!LocalRange"
    }


def test_data_validation_enforcement_is_compared(tmp_path):
    """A rule downgraded from "stop" to "information" lets anything through."""
    def build(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        validation = DataValidation(type="whole", operator="between",
                                    formula1="1", formula2="100")
        validation.errorStyle = "stop"
        validation.showErrorMessage = True
        ws.add_data_validation(validation)
        validation.add("B2:B4")
        wb.save(path)

    def defang(wb):
        validation = wb["Sheet1"].data_validations.dataValidation[0]
        validation.errorStyle = "information"
        validation.showErrorMessage = False

    baseline, candidate = _edited(tmp_path, build, defang)
    assert _attributes(diff(baseline, candidate))["data_validation"] == {
        "Sheet1!B2:B4"
    }


def test_conditional_format_styling_is_compared(tmp_path):
    """The rule still fires; it just stopped highlighting anything."""
    def build_with(color, path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.conditional_formatting.add("B2:B4", CellIsRule(
            operator="greaterThan", formula=["8"],
            fill=PatternFill(start_color=color, end_color=color, fill_type="solid")))
        wb.save(path)

    baseline = tmp_path / "base.xlsx"
    candidate = tmp_path / "cand.xlsx"
    build_with("FFC7CE", baseline)
    build_with("FFFFFF", candidate)
    assert _attributes(diff(baseline, candidate))["conditional_formatting"] == {
        "Sheet1!B2:B4"
    }


def test_rich_text_run_formatting_is_compared(tmp_path):
    """One run turned white inside a cell whose text never changed."""
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont

    def build_with(font, path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = CellRichText([TextBlock(font, "APPROVED"), " by finance"])
        wb.save(path)

    baseline = tmp_path / "base.xlsx"
    candidate = tmp_path / "cand.xlsx"
    build_with(InlineFont(b=True), baseline)
    build_with(InlineFont(b=True, color="FFFFFF"), candidate)

    changes = _attributes(diff(baseline, candidate))
    assert changes["format"] == {"Sheet1!A1"}
    assert "value" not in changes


def test_sheet_settings_are_compared(tmp_path):
    def build(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "x"
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:B3"
        ws.sheet_properties.tabColor = "FF0000"
        wb.save(path)

    def loosen(wb):
        ws = wb["Sheet1"]
        ws.freeze_panes = None
        ws.auto_filter.ref = None
        ws.sheet_properties.tabColor = "00FF00"

    baseline, candidate = _edited(tmp_path, build, loosen)
    assert _attributes(diff(baseline, candidate))["sheet_settings"] == {
        "Sheet1#settings"
    }


def test_workbook_calculation_mode_is_compared(tmp_path):
    """Manual calculation leaves every formula showing a stale answer."""
    baseline, candidate = _edited(
        tmp_path, _two_rows,
        lambda wb: setattr(wb.calculation, "calcMode", "manual"))
    assert _attributes(diff(baseline, candidate))["sheet_settings"] == {
        "workbook#settings"
    }


def test_result_json_roundtrip(fixtures):
    import json

    from templategate.reporters import render_json

    result = check(fixtures["excel_baseline"], fixtures["excel_bad"],
                   fixtures["excel_policy"])
    data = json.loads(render_json(result))
    assert data["tool"] == "templategate"
    assert data["passed"] is False
    assert data["summary"]["violations"] == len(result.violations)

"""Files written by tools other than openpyxl.

An .xlsx is a format, not a library, and the writers disagree about what to
put in it: SheetJS omits an element openpyxl assumes is present, ExcelJS
resets the workbook's default font while stamping the real one onto every
cell.  Neither is wrong, and neither is an edit — but one of them used to
make the entire file unreadable, and the other buried the report in noise.
"""

import re
import zipfile

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font

from templategate import check, diff, snapshot
from templategate.api import diff_report
from templategate.cli import main
from templategate.core.policy import parse_policy


def _rewrite(source, destination, part, transform):
    from generate import rewrite_zip

    with zipfile.ZipFile(source) as zf:
        patched = transform(zf.read(part).decode("utf-8"))
    rewrite_zip(source, destination, add={part: patched.encode("utf-8")})
    return destination


def _book(path, *, font=None, values=(("B3", 10), ("D7", "=SUM(D3:D5)"))):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for coord, value in values:
        ws[coord] = value
        if font is not None:
            ws[coord].font = font
    wb.save(path)
    return path


def _without_calc_properties(source, destination):
    """A workbook with no <calcPr>, the way SheetJS writes one."""
    return _rewrite(source, destination, "xl/workbook.xml",
                    lambda xml: re.sub(r"<calcPr[^>]*/>", "", xml))


# --- a missing optional element must not blind the whole reader ----------

def test_a_workbook_without_calc_properties_is_read_in_full(tmp_path):
    plain = _book(tmp_path / "plain.xlsx")
    book = _without_calc_properties(plain, tmp_path / "b.xlsx")
    with zipfile.ZipFile(book) as zf:
        assert b"<calcPr" not in zf.read("xl/workbook.xml")

    snap = snapshot(book)
    assert snap.get("degraded") is None
    assert set(snap["sheets"]["Sheet1"]["cells"]) == {"B3", "D7"}


def test_an_edit_is_visible_in_a_workbook_without_calc_properties(tmp_path):
    """The whole file used to be invisible because of one absent element."""
    baseline = _without_calc_properties(
        _book(tmp_path / "p1.xlsx"), tmp_path / "b.xlsx")
    candidate = _without_calc_properties(
        _book(tmp_path / "p2.xlsx", values=(("B3", 25), ("D7", 9999))),
        tmp_path / "c.xlsx")

    changes = {(c.location, c.attribute) for c in diff(baseline, candidate)}
    assert ("Sheet1!B3", "value") in changes
    assert ("Sheet1!D7", "formula") in changes


def test_an_absent_calc_mode_reads_as_auto(tmp_path):
    """Absent means the schema default, so it must not differ from "auto"."""
    plain = _book(tmp_path / "plain.xlsx")
    stripped = _without_calc_properties(plain, tmp_path / "b.xlsx")
    assert dict(snapshot(stripped)["settings"])["calc_mode"] == "auto"
    assert dict(snapshot(plain)["settings"])["calc_mode"] == "auto"
    assert diff(plain, stripped) == []


# --- diff must never say "no changes" about a file it could not read -----

@pytest.fixture()
def damaged(fixtures, tmp_path):
    from generate import rewrite_zip

    baseline = tmp_path / "base.docx"
    candidate = tmp_path / "cand.docx"
    rewrite_zip(fixtures["word_baseline"], baseline, add={})
    rewrite_zip(baseline, candidate, drop=("word/styles.xml",))
    return baseline, candidate


def test_diff_reports_the_damage_and_exits_two(damaged, capsys):
    baseline, candidate = damaged
    code = main(["diff", "--baseline", str(baseline), "--candidate", str(candidate)])
    captured = capsys.readouterr()
    assert code == 2
    assert "the candidate document is damaged" in captured.err
    assert "only its package parts could be compared" in captured.err
    assert "no changes" not in captured.out


def test_diff_never_says_no_changes_about_an_unreadable_file(fixtures, tmp_path,
                                                             capsys):
    """Silence here reads as "your edit was safe", which would be a lie."""
    from generate import rewrite_zip

    baseline = tmp_path / "base.docx"
    rewrite_zip(fixtures["word_baseline"], baseline, add={})
    candidate = tmp_path / "cand.docx"
    rewrite_zip(baseline, candidate, drop=("word/settings.xml",))

    code = main(["diff", "--baseline", str(baseline), "--candidate", str(candidate)])
    out = capsys.readouterr().out
    if code == 2:
        assert out.strip() != "no changes"


def test_a_healthy_diff_still_exits_zero(fixtures, capsys):
    code = main(["diff", "--baseline", str(fixtures["word_baseline"]),
                 "--candidate", str(fixtures["word_baseline"])])
    assert code == 0
    assert capsys.readouterr().out.strip() == "no changes"


def test_diff_report_hands_back_what_could_not_be_read(damaged):
    baseline, candidate = damaged
    changes, degraded = diff_report(baseline, candidate)
    assert "candidate" in degraded
    assert changes  # the parts that survived were still compared


# --- one writer's default is another writer's explicit value -------------

def _with_default_font(path, name):
    """A workbook whose *default* font is the given one."""
    return _rewrite(path, path.with_name("d_" + path.name), "xl/styles.xml",
                    lambda xml: xml.replace('<name val="Calibri"/>',
                                            f'<name val="{name}"/>', 1))


def test_a_cell_stating_the_font_it_used_to_inherit_is_silent(tmp_path):
    """ExcelJS resets the workbook default and stamps the real font on cells.

    The page renders identically; only who spells the font out has changed.
    """
    inheriting = _with_default_font(_book(tmp_path / "b.xlsx"), "游ゴシック")
    stating = _book(tmp_path / "c.xlsx", font=Font(name="游ゴシック"))

    fonts = [c for c in diff(inheriting, stating)
             if c.attribute == "format" and "font.name" in (c.detail or "")]
    assert fonts == []


def test_the_default_font_difference_is_still_reported_once(tmp_path):
    inheriting = _with_default_font(_book(tmp_path / "b.xlsx"), "游ゴシック")
    stating = _book(tmp_path / "c.xlsx", font=Font(name="游ゴシック"))
    settings = [c for c in diff(inheriting, stating)
                if c.location == "workbook#settings"]
    assert len(settings) == 1
    assert "游ゴシック -> Calibri" in settings[0].detail


def test_two_locales_materializing_their_own_default_stay_silent(tmp_path):
    """Both sides inherit; only the workbook default differs."""
    calibri = _book(tmp_path / "b.xlsx")
    japanese = _with_default_font(_book(tmp_path / "c.xlsx"), "ＭＳ Ｐゴシック")
    per_cell = [c for c in diff(calibri, japanese) if c.attribute == "format"]
    assert per_cell == []


def test_a_real_font_change_survives_all_of_that(tmp_path):
    """The normalization must not swallow a font someone actually changed."""
    baseline = _book(tmp_path / "b.xlsx", font=Font(name="游ゴシック"))
    candidate = _book(tmp_path / "c.xlsx", font=Font(name="Papyrus"))
    fonts = [c for c in diff(baseline, candidate) if c.attribute == "format"]
    assert fonts
    assert "font.name" in fonts[0].detail

    strict = parse_policy({"target": "excel",
                           "protect": [{"selector": "*", "attributes": ["*"]}]})
    assert not check(baseline, candidate, strict).passed


def test_an_absent_zoom_is_one_hundred_percent(tmp_path):
    plain = _book(tmp_path / "b.xlsx")
    assert dict(snapshot(plain)["sheets"]["Sheet1"]["settings"])["zoom"] == 100

    spelled_out = _rewrite(plain, tmp_path / "c.xlsx", "xl/worksheets/sheet1.xml",
                           lambda xml: xml.replace("<sheetView ",
                                                   '<sheetView zoomScale="100" ', 1))
    assert diff(plain, spelled_out) == []

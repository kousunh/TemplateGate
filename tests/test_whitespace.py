"""Pretty-printed worksheet XML must not black out cell comparison.

Nothing in OOXML says a worksheet has to be written on one line, and plenty
of things produce indented XML: hand-rolled writers, XSLT pipelines, and the
unzip/format/rezip round trip people reach for when debugging a package.
Excel opens all of them.

openpyxl's ``data_only`` load was the one reader in the pipeline that did
not — it fails with ``int('\\n')`` on whitespace between tags — and because a
failed load degraded the whole snapshot to package parts, one cosmetic
difference made an entire class of writers un-checkable at cell level.  The
gate now reads cached formula results straight from the worksheet XML and
does not perform that load at all.

The pair of properties this file pins: formatting alone is never a change,
and damage inside a pretty-printed file is still caught.
"""

import re
import zipfile

import pytest
from openpyxl import Workbook, load_workbook

from templategate import diff, snapshot

SHEET_PART = "xl/worksheets/sheet1.xml"

# The four shapes a formatter is likely to leave behind.
STYLES = {
    "newline": "\n",
    "newline_and_spaces": "\n    ",
    "tab": "\t",
    "single_space": " ",
}


def _workbook(path, total="=B1+B2"):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "alpha"
    ws["A2"] = "beta"
    ws["B1"] = 10
    ws["B2"] = 32
    ws["B3"] = total
    ws["B4"] = "=B1*2"
    wb.save(path)
    return path


def _with_cached_results(source, destination, value="42"):
    """Store an answer for B3 and an *empty* one for B4.

    The empty element is the whole point.  A writer that emits ``<v></v>``
    for a formula it did not evaluate produces ``<v>\\n</v>`` once the file is
    indented, and openpyxl's data_only load calls int() on that newline.  A
    fixture whose formulas all carry real numbers never reproduces the bug.
    """
    from generate import rewrite_zip

    with zipfile.ZipFile(source) as zf:
        xml = zf.read(SHEET_PART).decode("utf-8")
    stored = iter([f"<v>{value}</v>", "<v></v>"])
    patched = re.sub(r"<f>[^<]*</f>",
                     lambda m: m.group(0) + next(stored, "<v></v>"), xml)
    assert patched != xml, "expected formulas to attach cached results to"
    rewrite_zip(source, destination, add={SHEET_PART: patched.encode("utf-8")})
    return destination


def _pretty(source, destination, separator):
    """Put whitespace between every pair of tags, as a formatter would."""
    from generate import rewrite_zip

    with zipfile.ZipFile(source) as zf:
        xml = zf.read(SHEET_PART).decode("utf-8")
    body = xml.split("?>", 1)
    declaration, rest = (body[0] + "?>", body[1]) if len(body) == 2 else ("", xml)
    spaced = rest.replace("><", ">" + separator + "<")
    rewrite_zip(source, destination,
                add={SHEET_PART: (declaration + spaced).encode("utf-8")})
    return destination


@pytest.fixture
def compact(tmp_path):
    return _with_cached_results(_workbook(tmp_path / "raw.xlsx"),
                                tmp_path / "compact.xlsx")


# --- the load that used to fail -----------------------------------------

@pytest.mark.parametrize("style", sorted(STYLES))
def test_openpyxl_still_cannot_read_these(compact, tmp_path, style):
    """The premise of the whole file.  If openpyxl ever fixes this, these
    tests stop proving anything and should be revisited rather than deleted."""
    spaced = _pretty(compact, tmp_path / f"{style}.xlsx", STYLES[style])
    with pytest.raises(Exception):
        load_workbook(spaced, data_only=True)


@pytest.mark.parametrize("style", sorted(STYLES))
def test_a_pretty_printed_workbook_is_compared_in_full(compact, tmp_path, style):
    spaced = _pretty(compact, tmp_path / f"{style}.xlsx", STYLES[style])
    taken = snapshot(spaced)

    assert "error" not in taken, "the snapshot degraded instead of reading"
    cells = taken["sheets"]["Sheet"]["cells"]
    assert set(cells) == {"A1", "A2", "B1", "B2", "B3", "B4"}


@pytest.mark.parametrize("style", sorted(STYLES))
def test_the_cached_result_survives_the_whitespace(compact, tmp_path, style):
    """The value openpyxl could not reach is exactly what we now read."""
    spaced = _pretty(compact, tmp_path / f"{style}.xlsx", STYLES[style])
    assert snapshot(spaced)["sheets"]["Sheet"]["cells"]["B3"]["value"] == 42


# --- formatting is not a change -----------------------------------------

@pytest.mark.parametrize("style", sorted(STYLES))
def test_reformatting_the_xml_is_silent(compact, tmp_path, style):
    """Same content, different whitespace.  If the part hashes saw the bytes
    rather than the structure, every one of these would be a false positive."""
    spaced = _pretty(compact, tmp_path / f"{style}.xlsx", STYLES[style])
    assert diff(compact, spaced) == []


def test_reformatting_between_two_styles_is_silent(compact, tmp_path):
    one = _pretty(compact, tmp_path / "one.xlsx", STYLES["tab"])
    other = _pretty(compact, tmp_path / "other.xlsx", STYLES["newline_and_spaces"])
    assert diff(one, other) == []


# --- damage is still caught ---------------------------------------------

@pytest.mark.parametrize("style", sorted(STYLES))
def test_a_destroyed_formula_is_caught_inside_a_pretty_printed_file(
        compact, tmp_path, style):
    spaced = _pretty(compact, tmp_path / f"{style}.xlsx", STYLES[style])
    wrecked = _pretty(
        _workbook(tmp_path / f"{style}_raw.xlsx", total=42),
        tmp_path / f"{style}_wrecked.xlsx", STYLES[style])

    changes = diff(spaced, wrecked)
    assert any(c.attribute == "formula" and c.location.endswith("B3")
               for c in changes), [c.attribute for c in changes]


def test_a_changed_value_is_caught_inside_a_pretty_printed_file(compact, tmp_path):
    spaced = _pretty(compact, tmp_path / "spaced.xlsx", STYLES["newline"])
    edited = _pretty(
        _with_cached_results(_workbook(tmp_path / "edit_raw.xlsx"),
                            tmp_path / "edit_cached.xlsx"),
        tmp_path / "edited.xlsx", STYLES["newline"])
    # Change a literal, not the formula, so only the value moves.
    from generate import rewrite_zip
    with zipfile.ZipFile(edited) as zf:
        xml = zf.read(SHEET_PART).decode("utf-8")
    rewrite_zip(edited, tmp_path / "final.xlsx",
                add={SHEET_PART: xml.replace(">32<", ">99<").encode("utf-8")})

    changes = diff(spaced, tmp_path / "final.xlsx")
    assert any(c.location.endswith("B2") for c in changes)


def test_a_cached_result_changing_is_visible_in_a_pretty_printed_file(
        compact, tmp_path):
    """A formula whose stored answer no longer matches is the signal that
    somebody edited the cache by hand."""
    other = _with_cached_results(_workbook(tmp_path / "other_raw.xlsx"),
                                tmp_path / "other.xlsx", value="999")
    spaced = _pretty(compact, tmp_path / "a.xlsx", STYLES["newline"])
    spaced_other = _pretty(other, tmp_path / "b.xlsx", STYLES["newline"])

    changes = diff(spaced, spaced_other)
    assert any(c.location.endswith("B3") for c in changes)


# --- the reader itself ---------------------------------------------------

def test_every_cached_type_reads_back_the_way_openpyxl_read_it(tmp_path):
    """Numbers, strings, booleans and errors each have their own encoding in
    the sheet XML, and getting one wrong would silently change a value."""
    from generate import rewrite_zip
    from templategate.excel.snapshot import _cached_values

    wb = Workbook()
    ws = wb.active
    ws["A1"] = 2
    ws["B1"] = "=A1+1"
    ws["B2"] = '=IF(A1>1,"yes","no")'
    ws["B3"] = "=A1>5"
    ws["B4"] = "=1/0"
    wb.save(tmp_path / "types_raw.xlsx")

    with zipfile.ZipFile(tmp_path / "types_raw.xlsx") as zf:
        xml = zf.read(SHEET_PART).decode("utf-8")
    for ref, kind, value in (("B1", None, "3"), ("B2", "str", "yes"),
                             ("B3", "b", "0"), ("B4", "e", "#DIV/0!")):
        attribute = f' t="{kind}"' if kind else ""
        xml = re.sub(r'<c r="%s"([^>]*)>' % ref,
                     lambda m, a=attribute: f'<c r="{ref}"{m.group(1)}{a}>', xml)
        xml = re.sub(r'(<c r="%s".*?<f>[^<]*</f>)' % ref,
                     r"\1<v>%s</v>" % value, xml)
    typed = tmp_path / "types.xlsx"
    rewrite_zip(tmp_path / "types_raw.xlsx", typed,
                add={SHEET_PART: xml.encode("utf-8")})

    grid = _cached_values(typed)["Sheet"]
    assert grid[(1, 2)] == 3
    assert grid[(2, 2)] == "yes"
    assert grid[(3, 2)] is False
    assert grid[(4, 2)] == "#DIV/0!"


def test_a_formula_with_no_stored_answer_reads_as_none(tmp_path):
    """openpyxl writes formulas without a cached result, and that is not the
    same as a cached result of zero."""
    from templategate.excel.snapshot import _cached_values

    plain = _workbook(tmp_path / "plain.xlsx")
    assert _cached_values(plain)["Sheet"][(3, 2)] is None


def test_a_date_formula_result_is_still_a_date(tmp_path):
    """openpyxl's data_only load turned a serial into a datetime; dropping
    that load must not turn dates back into numbers."""
    import datetime

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "=TODAY()"
    ws["A1"].number_format = "yyyy-mm-dd"
    wb.save(tmp_path / "date_raw.xlsx")
    dated = _with_cached_results(tmp_path / "date_raw.xlsx",
                                tmp_path / "date.xlsx", value="45000")

    value = snapshot(dated)["sheets"]["Sheet"]["cells"]["A1"]["value"]
    assert isinstance(value, (datetime.datetime, datetime.date, str))
    assert "2023-03-15" in str(value)


def test_a_number_padded_with_whitespace_is_still_that_number(tmp_path):
    """A thorough formatter indents the contents of <v> too, not just the
    space between tags."""
    from generate import rewrite_zip
    from templategate.excel.snapshot import _cached_values

    cached = _with_cached_results(_workbook(tmp_path / "pad_raw.xlsx"),
                                  tmp_path / "pad_cached.xlsx")
    with zipfile.ZipFile(cached) as zf:
        xml = zf.read(SHEET_PART).decode("utf-8")
    padded = tmp_path / "padded.xlsx"
    rewrite_zip(cached, padded,
                add={SHEET_PART: xml.replace("<v>42</v>",
                                             "<v>\n    42\n  </v>").encode("utf-8")})

    assert _cached_values(padded)["Sheet"][(3, 2)] == 42


def test_an_empty_stored_answer_is_none_not_zero(tmp_path):
    """<v></v> means "no answer stored", and reading it as 0 would invent a
    value the file never claimed."""
    from templategate.excel.snapshot import _cached_values

    cached = _with_cached_results(_workbook(tmp_path / "empty_raw.xlsx"),
                                  tmp_path / "empty_cached.xlsx")
    spaced = _pretty(cached, tmp_path / "empty_spaced.xlsx", STYLES["newline"])

    assert _cached_values(spaced)["Sheet"][(4, 2)] is None

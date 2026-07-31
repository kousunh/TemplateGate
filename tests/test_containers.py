"""Containers that cannot be compared, names that are not what they look like,
and geometry that must not cost anything.
"""

import time
import unicodedata
import zipfile

import pytest
from openpyxl import Workbook

from templategate import check, diff, snapshot
from templategate.api import DocumentError
from templategate.cli import main
from templategate.core.package import package_problem
from templategate.core.policy import parse_policy
from templategate.core.selector import match_selector, quote_sheet


def _workbook(path, values, title="Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for coord, value in values.items():
        ws[coord] = value
    wb.save(path)
    return path


def _duplicate(source, destination, part, second):
    """A zip holding two members with the same name — legal, and unreadable."""
    with zipfile.ZipFile(source) as zin:
        items = [(i.filename, zin.read(i.filename)) for i in zin.infolist()]
    with zipfile.ZipFile(destination, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in items:
            zout.writestr(name, data)
        zout.writestr(part, second)
    return destination


# --- D2: an ambiguous container is refused, not guessed at ---------------

def test_duplicate_part_is_refused(tmp_path):
    baseline = _workbook(tmp_path / "base.xlsx", {"A1": "keep"})
    candidate = _duplicate(baseline, tmp_path / "cand.xlsx",
                           "xl/worksheets/sheet1.xml", b"<worksheet/>")
    problem = package_problem(candidate)
    assert problem is not None
    assert "2 copies of xl/worksheets/sheet1.xml" in problem
    with pytest.raises(DocumentError):
        snapshot(candidate)


def test_duplicate_part_exits_two_not_one(tmp_path, capsys):
    """It is not a damaged document; it is a document nobody can pin down."""
    baseline = _workbook(tmp_path / "base.xlsx", {"A1": "keep"})
    candidate = _duplicate(baseline, tmp_path / "cand.xlsx",
                           "xl/worksheets/sheet1.xml", b"<worksheet/>")
    policy = tmp_path / "p.yaml"
    policy.write_text("version: 1\ntarget: excel\n", encoding="utf-8")
    code = main(["check", "--baseline", str(baseline),
                 "--candidate", str(candidate), "--policy", str(policy)])
    assert code == 2
    assert "copies of" in capsys.readouterr().err


@pytest.mark.parametrize("name", [
    "../../evil.txt",
    "/etc/passwd",
    "..\\..\\evil.txt",
    "C:/windows/system32/evil.dll",
])
def test_part_escaping_the_package_root_is_refused(tmp_path, name):
    baseline = _workbook(tmp_path / "base.xlsx", {"A1": "keep"})
    candidate = _duplicate(baseline, tmp_path / "cand.xlsx", name, b"payload")
    problem = package_problem(candidate)
    assert problem is not None and "outside the package root" in problem
    with pytest.raises(DocumentError):
        snapshot(candidate)


def test_an_ordinary_workbook_has_no_problem(fixtures):
    assert package_problem(fixtures["excel_baseline"]) is None
    assert package_problem(fixtures["word_baseline"]) is None


def test_a_non_container_is_left_to_the_normal_error_path(tmp_path):
    broken = tmp_path / "broken.xlsx"
    broken.write_bytes(b"not a zip at all")
    assert package_problem(broken) is None


# --- D3: cost tracks content, not geometry -------------------------------

def test_a_far_corner_cell_does_not_walk_the_whole_grid(tmp_path):
    """A1 plus XFD1048576 declares a grid of 17 billion cells."""
    baseline = _workbook(tmp_path / "base.xlsx",
                         {"A1": "top left", "XFD1048576": "bottom right"})
    candidate = _workbook(tmp_path / "cand.xlsx",
                          {"A1": "top left", "XFD1048576": "TAMPERED"})

    started = time.monotonic()
    changes = diff(baseline, candidate)
    elapsed = time.monotonic() - started

    assert [(c.location, c.attribute) for c in changes] == [
        ("Sheet1!XFD1048576", "value")
    ]
    assert elapsed < 20, f"took {elapsed:.1f}s; cost is tracking geometry again"


def test_only_populated_cells_are_snapshotted(tmp_path):
    baseline = _workbook(tmp_path / "base.xlsx",
                         {"A1": "top left", "XFD1048576": "bottom right"})
    cells = snapshot(baseline)["sheets"]["Sheet1"]["cells"]
    assert set(cells) == {"A1", "XFD1048576"}


# --- D4: two spellings of one name are one name --------------------------

NFC_NAME = unicodedata.normalize("NFC", "Résumé")
NFD_NAME = unicodedata.normalize("NFD", "Résumé")


def test_the_two_spellings_really_do_differ():
    assert NFC_NAME != NFD_NAME
    assert len(NFC_NAME) != len(NFD_NAME)


def test_selector_matches_across_normal_forms():
    assert match_selector(f"{NFC_NAME}!A1", f"{NFD_NAME}!A1")
    assert match_selector(f"{NFD_NAME}!A1", f"{NFC_NAME}!A1")
    assert match_selector(NFC_NAME, f"{NFD_NAME}!B2")


def test_emitted_locations_are_composed(tmp_path):
    assert quote_sheet(NFD_NAME) == NFC_NAME
    baseline = _workbook(tmp_path / "base.xlsx", {"A1": "SECRET"}, title=NFD_NAME)
    candidate = _workbook(tmp_path / "cand.xlsx", {"A1": "LEAKED"}, title=NFD_NAME)
    locations = {c.location for c in diff(baseline, candidate)}
    assert locations == {f"{NFC_NAME}!A1"}


@pytest.mark.parametrize("policy_form,file_form", [
    (NFC_NAME, NFD_NAME),
    (NFD_NAME, NFC_NAME),
])
def test_a_protect_rule_reaches_the_sheet_however_it_was_typed(
        tmp_path, policy_form, file_form):
    """A rule that silently matches nothing is worse than no rule at all."""
    baseline = _workbook(tmp_path / "base.xlsx", {"A1": "SECRET"}, title=file_form)
    candidate = _workbook(tmp_path / "cand.xlsx", {"A1": "LEAKED"}, title=file_form)
    policy = parse_policy({
        "target": "excel",
        "allow": [{"selector": "*", "attributes": ["*"]}],
        "protect": [{"selector": f"{policy_form}!A1", "attributes": ["value"]}],
    })
    result = check(baseline, candidate, policy)
    assert not result.passed
    assert result.violations[0].rule == "protected"


# --- D5: suppressed error flags --------------------------------------------

def test_suppressed_error_indicators_are_compared(tmp_path):
    """Turning off the warning triangles is how a broken formula looks fine."""
    from generate import rewrite_zip

    baseline = _workbook(tmp_path / "base.xlsx", {"A1": "1", "A2": "=A1+1"})
    with zipfile.ZipFile(baseline) as zf:
        sheet = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
    ignored = ('<ignoredErrors><ignoredError sqref="A1:A9" '
               'numberStoredAsText="1"/></ignoredErrors>')
    candidate = tmp_path / "cand.xlsx"
    rewrite_zip(baseline, candidate, add={
        "xl/worksheets/sheet1.xml":
            sheet.replace("</worksheet>", ignored + "</worksheet>").encode()})

    changes = [c for c in diff(baseline, candidate)
               if c.attribute == "sheet_settings"]
    assert changes
    assert "ignored_errors" in changes[0].detail


# --- D6: a format diff names what moved ----------------------------------

def test_a_colour_change_reads_as_a_colour_change(tmp_path):
    from openpyxl.styles import Font

    def build(path, color):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "total"
        ws["A1"].font = Font(color=color)
        wb.save(path)

    baseline = build(tmp_path / "base.xlsx", "FF000000") or tmp_path / "base.xlsx"
    candidate = build(tmp_path / "cand.xlsx", "FFFFFFFF") or tmp_path / "cand.xlsx"
    change = [c for c in diff(baseline, candidate) if c.attribute == "format"][0]
    assert change.detail == "cell format changed: font.color FF000000 -> FFFFFFFF"
    assert change.old == {"font.color": "FF000000"}


def test_no_openpyxl_descriptor_leaks_into_a_report(tmp_path):
    from openpyxl.styles import Font

    for name, color in (("base", "FF000000"), ("cand", "FFFFFFFF")):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "x"
        ws["A1"].font = Font(color=color)
        wb.save(tmp_path / f"{name}.xlsx")

    rendered = repr(diff(tmp_path / "base.xlsx", tmp_path / "cand.xlsx"))
    for noise in ("Integer[", "Typed[", "object at 0x", "Descriptor"):
        assert noise not in rendered


def test_a_cell_gaining_a_style_reports_only_what_it_set(tmp_path):
    from openpyxl.styles import Font

    plain = _workbook(tmp_path / "base.xlsx", {"A1": "x"})
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "x"
    ws["A1"].font = Font(bold=True)
    wb.save(tmp_path / "cand.xlsx")

    change = [c for c in diff(plain, tmp_path / "cand.xlsx")
              if c.attribute == "format"][0]
    assert list(change.new) == ["font.bold"]
    assert "none -> none" not in change.detail


# --- a Word document with more than one body -----------------------------

def _reshaped_document(source, destination, transform):
    from generate import rewrite_zip

    with zipfile.ZipFile(source) as zf:
        xml = zf.read("word/document.xml").decode("utf-8")
    rewrite_zip(source, destination,
                add={"word/document.xml": transform(xml).encode("utf-8")})
    return destination


SECOND_BODY = ("</w:body><w:body><w:p><w:r>"
               "<w:t>Payment due: USD 25 only.</w:t></w:r></w:p></w:body>")


@pytest.fixture
def two_bodies(fixtures, tmp_path):
    return _reshaped_document(
        fixtures["word_baseline"], tmp_path / "two_bodies.docx",
        lambda xml: xml.replace("</w:body>", SECOND_BODY, 1))


def test_a_second_body_hides_content_from_every_reader_we_have(two_bodies):
    """The premise: this is why the file has to be refused rather than read.

    Both the XML walker and python-docx stop at the first body, so the second
    body's text is invisible to the gate while another reader may show it.
    """
    import xml.etree.ElementTree as ElementTree

    import docx

    from templategate.word.content import W

    with zipfile.ZipFile(two_bodies) as zf:
        document = zf.read("word/document.xml")
    assert b"Payment due" in document, "the fixture did not add the second body"

    # Every reader in play resolves the body with find(), which returns the
    # first match and never looks for a second.
    root = ElementTree.fromstring(document)
    first = root.find(W + "body")
    assert "Payment due" not in "".join(
        node.text or "" for node in first.iter(W + "t"))

    assert "Payment due" not in "\n".join(
        p.text for p in docx.Document(str(two_bodies)).paragraphs)


def test_a_document_with_two_bodies_is_refused(two_bodies):
    with pytest.raises(DocumentError) as caught:
        snapshot(two_bodies)
    assert "2 <w:body>" in str(caught.value)


def test_the_refusal_names_the_part_and_the_reason(two_bodies):
    assert "word/document.xml" in package_problem(two_bodies)
    assert "different readers" in package_problem(two_bodies)


def test_two_bodies_exit_two_not_one(fixtures, two_bodies, capsys):
    """A file nobody can pin down is a tool-level refusal, not a policy FAIL."""
    code = main(["check", "--baseline", str(fixtures["word_baseline"]),
                 "--candidate", str(two_bodies),
                 "--policy", str(fixtures["word_policy"])])
    assert code == 2
    error = capsys.readouterr().err
    assert "Traceback" not in error
    assert "<w:body>" in error


def test_a_second_body_is_refused_whatever_the_namespace_prefix(fixtures, tmp_path):
    """Counting the tag as a string would miss a document that binds the
    wordprocessingml namespace to a different prefix."""
    def rebind(xml):
        with_second = xml.replace("</w:body>", SECOND_BODY, 1)
        return with_second.replace(
            'xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"',
            'xmlns:zz="http://schemas.openxmlformats.org/wordprocessingml/2006/main"'
        ).replace("<w:", "<zz:").replace("</w:", "</zz:").replace('w:', 'zz:')

    renamed = _reshaped_document(fixtures["word_baseline"],
                                 tmp_path / "prefixed.docx", rebind)
    assert package_problem(renamed) is not None


def test_text_after_the_body_is_refused(fixtures, tmp_path):
    stray = _reshaped_document(
        fixtures["word_baseline"], tmp_path / "tail.docx",
        lambda xml: xml.replace("</w:body>", "</w:body>Payment due: USD 25.", 1))
    with pytest.raises(DocumentError) as caught:
        snapshot(stray)
    assert "after </w:body>" in str(caught.value)


def test_whitespace_after_the_body_is_not_a_problem(fixtures, tmp_path):
    """Indented XML puts a newline there, and that says nothing."""
    spaced = _reshaped_document(
        fixtures["word_baseline"], tmp_path / "spaced.docx",
        lambda xml: xml.replace("</w:body>", "</w:body>\n  ", 1))
    assert package_problem(spaced) is None
    assert "degraded" not in snapshot(spaced)


def test_a_normal_document_is_untouched(fixtures):
    assert package_problem(fixtures["word_baseline"]) is None
    assert "degraded" not in snapshot(fixtures["word_baseline"])


def test_a_normal_document_still_passes_its_policy(fixtures):
    result = check(fixtures["word_baseline"], fixtures["word_good"],
                   fixtures["word_policy"])
    assert result.passed, [v.message for v in result.violations]

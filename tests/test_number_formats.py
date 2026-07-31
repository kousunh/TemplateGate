"""Formats Excel assigns outside the English-language builtins.

Number-format ids 27-36 and 50-58 are reserved by the spec for East Asian
locales, and openpyxl resolves none of them — it reports "General", which is
also what an unformatted cell reports.  On a Japanese quotation that made
「2026年7月1日」 indistinguishable from a bare number, so switching a date to
the imperial era, or to a time, passed the gate in silence.

The same unreadability leaked into the *value*: openpyxl decides whether to
turn a serial number into a datetime by reading the format string, so
changing only the format changed the value too.
"""

import zipfile

import pytest
from openpyxl import Workbook

from templategate import check, diff, snapshot
from templategate.core.policy import parse_policy

STRICT = parse_policy({
    "target": "excel",
    "protect": [{"selector": "*", "attributes": ["*"]}],
})


def _dated(path, format_id):
    """A workbook whose date cell carries a given builtin number-format id."""
    from generate import rewrite_zip

    plain = path.with_name(path.name + ".plain.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "見積書"
    ws["A1"] = "発行日"
    ws["F3"] = 46204  # 2026-07-01 as an Excel serial
    # A builtin openpyxl *does* know, so its id appears verbatim in styles.xml
    # exactly once and can be repointed at the locale builtin under test.
    ws["F3"].number_format = "mm-dd-yy"
    wb.save(plain)

    if format_id == 14:
        return plain

    with zipfile.ZipFile(plain) as zf:
        styles = zf.read("xl/styles.xml").decode("utf-8")
    assert styles.count('numFmtId="14"') == 1, styles
    styles = styles.replace('numFmtId="14"', f'numFmtId="{format_id}"')
    rewrite_zip(plain, path, add={"xl/styles.xml": styles.encode("utf-8")})
    return path


def _formats(path):
    return snapshot(path)["sheets"]["見積書"]["cells"]["F3"]


def _only(changes, attribute):
    return [c for c in changes if c.attribute == attribute]


def test_a_locale_builtin_resolves_to_its_documented_code(tmp_path):
    """openpyxl calls it General; the id says which format it really is."""
    cell = _formats(_dated(tmp_path / "b.xlsx", 31))
    assert dict(cell["format"])["numfmt"] == 'yyyy"年"m"月"d"日"'


def test_western_year_to_imperial_era_is_caught(tmp_path):
    """31 -> 27 turns 2026年7月1日 into the imperial era; both read General."""
    baseline = _dated(tmp_path / "b.xlsx", 31)
    candidate = _dated(tmp_path / "c.xlsx", 27)
    change = _only(diff(baseline, candidate), "format")[0]
    assert change.detail == ('cell format changed: numfmt '
                             'yyyy"年"m"月"d"日" -> ge.m.d')
    assert not check(baseline, candidate, STRICT).passed


def test_a_date_turned_into_a_time_is_caught(tmp_path):
    """31 -> 32 makes the issue date render as 0時00分."""
    baseline = _dated(tmp_path / "b.xlsx", 31)
    candidate = _dated(tmp_path / "c.xlsx", 32)
    change = _only(diff(baseline, candidate), "format")[0]
    assert 'h"時"mm"分"' in change.detail
    assert not check(baseline, candidate, STRICT).passed


def test_the_same_format_twice_is_silent(tmp_path):
    baseline = _dated(tmp_path / "b.xlsx", 31)
    candidate = _dated(tmp_path / "c.xlsx", 31)
    assert diff(baseline, candidate) == []


def test_changing_only_the_format_does_not_change_the_value(tmp_path):
    """The value used to flip between a serial number and an ISO datetime."""
    baseline = _dated(tmp_path / "b.xlsx", 31)
    candidate = _dated(tmp_path / "c.xlsx", 14)
    changes = diff(baseline, candidate)
    assert [c.attribute for c in changes] == ["format"]
    assert _formats(baseline)["value"] == _formats(candidate)["value"]


def test_a_readable_format_reads_as_itself(tmp_path):
    baseline = _dated(tmp_path / "b.xlsx", 31)
    candidate = _dated(tmp_path / "c.xlsx", 14)
    change = _only(diff(baseline, candidate), "format")[0]
    assert change.detail.endswith("-> mm-dd-yy")
    assert "none" not in change.detail


def test_the_id_and_the_written_out_code_are_the_same_format(tmp_path):
    """Excel stores 「2026年7月1日」 as builtin 31; openpyxl writes the code out.

    One is shorthand for the other, and a library that converts between them
    on save has not changed how a single cell renders.
    """
    from generate import rewrite_zip

    as_builtin = _dated(tmp_path / "b.xlsx", 31)
    with zipfile.ZipFile(as_builtin) as zf:
        styles = zf.read("xl/styles.xml").decode("utf-8")
    written_out = tmp_path / "c.xlsx"
    rewrite_zip(as_builtin, written_out, add={
        "xl/styles.xml": styles.replace(
            "<numFmts count=\"0\"/>",
            '<numFmts count="1"><numFmt numFmtId="176" '
            'formatCode="yyyy&quot;年&quot;m&quot;月&quot;d&quot;日&quot;"/></numFmts>'
        ).replace('numFmtId="31"', 'numFmtId="176"').encode("utf-8")})
    assert diff(as_builtin, written_out) == []


def test_an_unknown_builtin_never_reads_as_no_format(tmp_path):
    """The dangerous case: a format we cannot resolve must say so."""
    cell = _formats(_dated(tmp_path / "b.xlsx", 23))
    assert dict(cell["format"])["numfmt"] == "builtin:23 (unresolved)"


def test_losing_a_format_reads_as_general(tmp_path):
    """An absent number format is General, not "none"."""
    from openpyxl.styles import Font

    def build(path, number_format):
        wb = Workbook()
        ws = wb.active
        ws.title = "見積書"
        ws["F3"] = 1000
        ws["F3"].number_format = number_format
        ws["F3"].font = Font(bold=True)
        wb.save(path)
        return path

    baseline = build(tmp_path / "b.xlsx", "#,##0")
    candidate = build(tmp_path / "c.xlsx", "General")
    change = _only(diff(baseline, candidate), "format")[0]
    assert change.detail == "cell format changed: numfmt #,##0 -> General"


def test_a_real_date_change_still_fails(tmp_path):
    """Normalizing the value must not make the value uncomparable."""
    baseline = _dated(tmp_path / "b.xlsx", 31)
    candidate = _dated(tmp_path / "c.xlsx", 31)
    from generate import rewrite_zip

    with zipfile.ZipFile(candidate) as zf:
        sheet = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
    moved = tmp_path / "d.xlsx"
    rewrite_zip(candidate, moved, add={
        "xl/worksheets/sheet1.xml":
            sheet.replace("<v>46204</v>", "<v>46205</v>").encode("utf-8")})
    assert _only(diff(baseline, moved), "value")

"""What the tool says when the policy, not the document, is wrong.

A gate fails closed, which means a misspelled policy is silently useless
rather than loudly broken: `protekt:` protects nothing while looking like it
protects everything.  These tests pin the messages that turn that into a
two-second fix.
"""

import json

import pytest
from openpyxl import Workbook

from templategate import check
from templategate.cli import main
from templategate.core.liveness import policy_warnings
from templategate.core.policy import (
    SAMPLE_POLICY_EXCEL,
    SAMPLE_POLICY_WORD,
    PolicyError,
    parse_policy,
)
from templategate.reporters import render_json, render_markdown, render_text


def _error(raw) -> str:
    with pytest.raises(PolicyError) as caught:
        parse_policy(raw)
    return str(caught.value)


# --- E1: what is ignored today must be rejected today --------------------

@pytest.mark.parametrize("raw,expected", [
    ({"allowed": []}, "unknown policy key 'allowed' — did you mean 'allow'?"),
    ({"protekt": []}, "unknown policy key 'protekt' — did you mean 'protect'?"),
    ({"target": "exel"}, "target: 'exel' is not a target — did you mean 'excel'?"),
    ({"mode": "reviewonly"},
     "mode: 'reviewonly' is not a mode — did you mean 'review_only'?"),
    ({"allow": [{"selector": "*", "attributes": ["valeu"]}]},
     "allow: unknown attribute 'valeu' — did you mean 'value'?"),
    ({"protect": [{"selector": "*", "attributes": ["formulas"]}]},
     "protect: unknown attribute 'formulas' — did you mean 'formula'?"),
    ({"allow": [{"selctor": "*"}]},
     "allow: unknown key 'selctor' — did you mean 'selector'?"),
    ({"structural": {"sheets": "stricct"}},
     "structural.sheets: 'stricct' is not a setting — did you mean 'strict'?"),
    ({"structural": {"sheat": "strict"}},
     "unknown structural key 'sheat' — did you mean 'sheets'?"),
    ({"semantic": {"moed": "off"}},
     "semantic: unknown key 'moed' — did you mean 'mode'?"),
    ({"semantic": {"mode": "reveiw"}},
     "semantic.mode: 'reveiw' is not a mode — did you mean 'review'?"),
])
def test_a_typo_is_named_and_corrected(raw, expected):
    assert _error(raw) == expected


def test_a_word_with_no_near_match_lists_the_valid_ones():
    message = _error({"structural": {"sheets": "sometimes"}})
    assert "valid: ignore, strict" in message


@pytest.mark.parametrize("raw", [
    {},
    {"version": 1, "target": "auto"},
    {"target": "excel", "allow": ["*"], "protect": [{"selector": "*"}]},
    {"target": "word", "mode": "page_extension",
     "allow": [{"selector": "body", "attributes": ["text", "*"]}]},
    {"target": "excel", "structural": {"sheets": "ignore", "links": "strict"}},
    {"target": "excel", "semantic": {"mode": "gate", "provider": "command",
                                     "command": "x", "model": "m",
                                     "checks": ["a"]}},
])
def test_everything_valid_still_parses(raw):
    assert parse_policy(raw) is not None


def test_both_starter_policies_survive_their_own_validation():
    """`templategate init` must not write a file the parser rejects."""
    import yaml

    for sample in (SAMPLE_POLICY_EXCEL, SAMPLE_POLICY_WORD):
        assert parse_policy(yaml.safe_load(sample)) is not None


def test_a_bad_policy_exits_two(tmp_path, fixtures, capsys):
    policy = tmp_path / "p.yaml"
    policy.write_text("version: 1\ntarget: excel\nallowed: []\n", encoding="utf-8")
    code = main(["check", "--baseline", str(fixtures["excel_baseline"]),
                 "--candidate", str(fixtures["excel_good"]), "--policy", str(policy)])
    assert code == 2
    assert "did you mean 'allow'" in capsys.readouterr().err


# --- E2: a rule that reaches nothing says so ------------------------------

def _sheet(path, title, values):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for coord, value in values.items():
        ws[coord] = value
    wb.save(path)
    return path


def test_a_misspelled_sheet_name_is_warned_about(tmp_path):
    baseline = _sheet(tmp_path / "b.xlsx", "Budget", {"B2": 1})
    candidate = _sheet(tmp_path / "c.xlsx", "Budget", {"B2": 2})
    policy = parse_policy({
        "target": "excel",
        "allow": [{"selector": "Sheet1!B2:B100", "attributes": ["value"]}],
    })
    result = check(baseline, candidate, policy)
    assert len(result.warnings) == 1
    warning = result.warnings[0]
    assert "allow rule 1" in warning
    assert "no sheet named 'Sheet1'" in warning
    assert "'Budget'" in warning


def test_a_warning_is_not_a_violation(tmp_path):
    """It is advice about the policy, not a judgement on the document."""
    baseline = _sheet(tmp_path / "b.xlsx", "Budget", {"B2": 1})
    policy = parse_policy({
        "target": "excel",
        "allow": [{"selector": "Nope!A1", "attributes": ["value"]}],
    })
    result = check(baseline, baseline, policy)
    assert result.warnings
    assert result.passed
    assert result.violations == []


@pytest.mark.parametrize("selector", [
    "*", "Budget", "Budget!A1:Z99", "Budget!Z90:Z99",
    "sheet:*", "name:*", "vba", "package#*", "package#charts:*",
    "workbook#settings",
])
def test_no_warning_when_the_answer_is_not_obvious(tmp_path, selector):
    """Silence beats a wrong warning; people stop reading noisy ones."""
    baseline = _sheet(tmp_path / "b.xlsx", "Budget", {"B2": 1})
    policy = parse_policy({
        "target": "excel", "protect": [{"selector": selector}],
    })
    assert policy_warnings(policy, __import__("templategate").snapshot(baseline)) == []


def test_a_paragraph_past_the_end_is_warned_about(fixtures):
    policy = parse_policy({
        "target": "word", "allow": [{"selector": "p90", "attributes": ["text"]}],
    })
    result = check(fixtures["word_baseline"], fixtures["word_good"], policy)
    assert len(result.warnings) == 1
    assert "4 body paragraphs" in result.warnings[0]


@pytest.mark.parametrize("selector", ["body", "p1-20", "p1", "table1", "section1"])
def test_word_selectors_that_do_reach_something_stay_quiet(fixtures, selector):
    policy = parse_policy({"target": "word", "protect": [{"selector": selector}]})
    result = check(fixtures["word_baseline"], fixtures["word_good"], policy)
    assert result.warnings == []


def test_warnings_reach_every_report(tmp_path):
    baseline = _sheet(tmp_path / "b.xlsx", "Budget", {"B2": 1})
    policy = parse_policy({
        "target": "excel", "protect": [{"selector": "Nope!A1"}],
    })
    result = check(baseline, baseline, policy)
    assert "Policy warnings" in render_text(result)
    assert "Policy warnings" in render_markdown(result)
    assert json.loads(render_json(result))["warnings"] == result.warnings


def test_warnings_are_printed_to_stderr(tmp_path, capsys):
    baseline = _sheet(tmp_path / "b.xlsx", "Budget", {"B2": 1})
    policy = tmp_path / "p.yaml"
    policy.write_text('version: 1\ntarget: excel\nprotect:\n  - selector: "Nope!A1"\n',
                      encoding="utf-8")
    main(["check", "--baseline", str(baseline), "--candidate", str(baseline),
          "--policy", str(policy)])
    assert "warning:" in capsys.readouterr().err


# --- E3: every flag explains itself ---------------------------------------

@pytest.mark.parametrize("command",
                         ["check", "diff", "snapshot", "init", "suggest"])
def test_every_subcommand_documents_its_flags(capsys, command):
    with pytest.raises(SystemExit):
        main([command, "--help"])
    out = capsys.readouterr().out
    assert "Examples:" in out
    assert "templategate " + command in out
    # No flag may be left with an empty help column.
    for line in out.splitlines():
        stripped = line.strip()
        if stripped.startswith("--") and " " not in stripped:
            pytest.fail(f"{command}: {stripped} has no help text")


def test_defaults_are_stated(capsys):
    with pytest.raises(SystemExit):
        main(["check", "--help"])
    assert "(default: text)" in capsys.readouterr().out


def test_exit_codes_are_explained(capsys):
    with pytest.raises(SystemExit):
        main(["--help"])
    out = capsys.readouterr().out
    assert "Exit codes:" in out


# --- E4: report nits ------------------------------------------------------

def test_a_violation_line_does_not_repeat_its_location(tmp_path):
    baseline = _sheet(tmp_path / "b.xlsx", "Budget", {"B2": 1})
    candidate = _sheet(tmp_path / "c.xlsx", "Budget", {"B2": 2})
    policy = parse_policy({
        "target": "excel", "protect": [{"selector": "*", "attributes": ["value"]}],
    })
    report = render_text(check(baseline, candidate, policy))
    line = [ln for ln in report.splitlines() if "Budget!B2" in ln][0]
    assert line.count("Budget!B2") == 1


def test_a_formula_replaced_by_its_own_answer_reads_as_one_story(tmp_path):
    def build(path, b4):
        wb = Workbook()
        ws = wb.active
        ws.title = "Budget"
        ws["B2"], ws["B3"] = 1000, 750
        ws["B4"] = b4
        wb.save(path)

    baseline = tmp_path / "b.xlsx"
    candidate = tmp_path / "c.xlsx"
    build(baseline, "=SUM(B2:B3)")
    build(candidate, 1750)

    from templategate import diff

    formula = [c for c in diff(baseline, candidate) if c.attribute == "formula"][0]
    assert formula.detail == (
        "formula replaced by a hardcoded value (=SUM(B2:B3) -> 1750)")


def test_no_python_repr_leaks_into_the_text_report(tmp_path):
    baseline = _sheet(tmp_path / "b.xlsx", "Budget", {"B2": 1, "C2": "keep"})
    candidate = _sheet(tmp_path / "c.xlsx", "Budget", {"B2": 2})
    policy = parse_policy({
        "target": "excel", "protect": [{"selector": "*", "attributes": ["*"]}],
    })
    report = render_text(check(baseline, candidate, policy))
    assert "None" not in report
    assert "{'" not in report and "': " not in report


def test_an_absent_value_reads_as_a_word(tmp_path):
    baseline = _sheet(tmp_path / "b.xlsx", "Budget", {"B2": "was here"})
    candidate = _sheet(tmp_path / "c.xlsx", "Budget", {"A1": "x"})
    policy = parse_policy({
        "target": "excel", "protect": [{"selector": "*", "attributes": ["*"]}],
    })
    report = render_text(check(baseline, candidate, policy))
    assert "was here -> none" in report


# --- E5: the starter policies are honest about their format ---------------

def test_excel_only_categories_stay_out_of_the_word_starter():
    """Word has no pivot table or drawing part family to protect."""
    for key in ("pivot_tables:", "drawings:", "sheets:", "defined_names:"):
        assert key not in SAMPLE_POLICY_WORD
    for key in ("pivot_tables:", "drawings:", "sheets:", "defined_names:"):
        assert key in SAMPLE_POLICY_EXCEL


def test_word_only_categories_stay_out_of_the_excel_starter():
    assert "tables:" in SAMPLE_POLICY_WORD
    assert "\n  tables:" not in SAMPLE_POLICY_EXCEL


def test_shared_categories_appear_in_both():
    for key in ("charts:", "comments:", "embedded:", "custom_xml:",
                "parts:", "links:", "images:"):
        assert key in SAMPLE_POLICY_EXCEL, key
        assert key in SAMPLE_POLICY_WORD, key

"""Drafting a policy from one edit you have already reviewed.

Writing a policy from a blank page is the step where TemplateGate loses
people: you cannot list what an agent is allowed to change until you have
watched it change something.  `templategate suggest` inverts that — show it
a baseline and a candidate you have checked by eye, and it writes the policy
that would have accepted that edit.

The rule the whole feature turns on is that a draft never allows anything it
could not explain.  Collateral it recognises (a cached formula result, a
chart's redrawn cache, a row height read back by the editor) is allowed with
the reason written next to it.  Anything else is listed in the header as
*not* allowed, so a draft generated from a compromised edit refuses to bless
the compromise.
"""

import re
import zipfile

import pytest
from openpyxl import Workbook

from templategate import check, snapshot
from templategate.core.policy import load_policy
from templategate.suggest import COLLATERAL, INTENDED, SUSPICIOUS, classify, draft


# --- a quote, shaped like the one from the practicality evaluation -------

def _quote(path):
    """Items, quantities, and subtotals that are formulas.

    The hidden helper column matters: real templates carry them, and the
    editor writes a width back onto them on every save.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "見積書"
    ws["B3"] = "御見積書"
    ws["F3"] = "2026-07-31"
    for row, (name, unit) in enumerate(
            [("設計", 50000), ("実装", 80000), ("試験", 30000)], start=10):
        ws[f"B{row}"] = name
        ws[f"C{row}"] = unit
        ws[f"D{row}"] = 1
        ws[f"E{row}"] = f"=C{row}*D{row}"
    ws["E14"] = "=SUM(E10:E12)"
    ws["E15"] = "=E14*0.1"
    ws["E16"] = "=E14+E15"
    ws.column_dimensions["H"].hidden = True
    wb.save(path)
    return path


def _with_cached_results(source, destination):
    """Give every formula a stored answer, the way Excel would.

    openpyxl never computes, so a baseline it wrote has no cached results
    and the most common piece of real collateral cannot appear.
    """
    from generate import rewrite_zip

    with zipfile.ZipFile(source) as zf:
        name = "xl/worksheets/sheet1.xml"
        xml = zf.read(name).decode("utf-8")
    patched = re.sub(r"(<f>[^<]*</f>)", r"\1<v>1</v>", xml)
    assert patched != xml, "no formulas were found to cache"
    rewrite_zip(source, destination, add={name: patched.encode("utf-8")})
    return destination


def _legit_edit(source, destination):
    """A quantity changes.  Saving drops every cached result."""
    from openpyxl import load_workbook

    wb = load_workbook(source)
    wb["見積書"]["D10"] = 3
    wb.save(destination)
    return destination


def _accident(source, destination):
    """The subtotal formula is overwritten with the number it showed."""
    from openpyxl import load_workbook

    wb = load_workbook(source)
    wb["見積書"]["E16"] = 264000
    wb.save(destination)
    return destination


@pytest.fixture
def quote(tmp_path):
    raw = _quote(tmp_path / "raw.xlsx")
    baseline = _with_cached_results(raw, tmp_path / "baseline.xlsx")
    return {
        "baseline": baseline,
        "legit": _legit_edit(baseline, tmp_path / "legit.xlsx"),
        "accident": _accident(baseline, tmp_path / "accident.xlsx"),
    }


# --- the headline promise ------------------------------------------------

def test_a_draft_accepts_the_edit_it_was_drawn_from(quote, tmp_path):
    policy = tmp_path / "draft.yaml"
    policy.write_text(draft(quote["baseline"], quote["legit"]), encoding="utf-8")

    result = check(quote["baseline"], quote["legit"], policy)
    assert result.passed, [v.message for v in result.violations]


def test_a_draft_still_catches_the_accident(quote, tmp_path):
    """The point of the draft is that it is narrow, not that it is quiet."""
    policy = tmp_path / "draft.yaml"
    policy.write_text(draft(quote["baseline"], quote["legit"]), encoding="utf-8")

    result = check(quote["baseline"], quote["accident"], policy)
    assert not result.passed
    assert any("E16" in v.change.location for v in result.violations)


def test_a_draft_is_a_policy_the_loader_accepts(quote, tmp_path):
    """Every selector it emits has to survive a round trip through parsing."""
    policy = tmp_path / "draft.yaml"
    policy.write_text(draft(quote["baseline"], quote["legit"]), encoding="utf-8")

    loaded = load_policy(policy)
    assert loaded.target == "excel"
    assert loaded.allow, "a draft with no allow rules would pass nothing"


def test_the_sheet_name_is_quoted_so_yaml_keeps_it(quote):
    text = draft(quote["baseline"], quote["legit"])
    assert '"見積書!D10"' in text


# --- what it refuses to do ----------------------------------------------

def test_a_destroyed_formula_is_never_allowed(quote, tmp_path):
    """Drafting from a bad edit must not launder it into a policy."""
    text = draft(quote["baseline"], quote["accident"])

    assert "NOT allowed by this draft" in text
    assert "E16" in text.split("allow:")[0], "the refusal belongs in the header"

    policy = tmp_path / "draft.yaml"
    policy.write_text(text, encoding="utf-8")
    result = check(quote["baseline"], quote["accident"], policy)
    assert not result.passed


def test_a_draft_from_a_bad_edit_says_it_will_not_pass(quote):
    text = draft(quote["baseline"], quote["accident"])
    assert "does not pass the edit it was generated from" in text


def test_formulas_stay_protected_even_though_none_changed(quote):
    text = draft(quote["baseline"], quote["legit"])
    protect = text.split("protect:")[1]
    assert "formula" in protect and "vba" in protect


def test_the_header_warns_against_an_agent_widening_its_own_policy(quote):
    text = draft(quote["baseline"], quote["legit"])
    assert "must not be the one that runs" in text


# --- classification ------------------------------------------------------

def test_a_cached_result_is_collateral_not_an_edit(quote):
    baseline = snapshot(quote["baseline"])
    from templategate import diff

    cached = [c for c in diff(quote["baseline"], quote["legit"])
              if c.location.endswith("E10")]
    assert cached, "expected the cached result of E10 to have been dropped"
    verdict, reason = classify(cached[0], baseline)
    assert verdict == COLLATERAL
    assert "cached" in reason


def test_a_width_on_a_hidden_column_is_collateral(quote):
    """The detail text reads "still hidden" — the classifier must not
    conclude from that word that the hidden state moved."""
    from templategate.core.model import ATTR_LAYOUT, Change

    change = Change(location="見積書!H:H", attribute=ATTR_LAYOUT,
                    old={"width": 0.0}, new={"width": 13.0},
                    detail="row or column changed: column width 0.0 -> 13.0"
                           " (still hidden)")
    verdict, _ = classify(change, {})
    assert verdict == COLLATERAL


def test_actually_hiding_a_column_is_not_collateral(quote):
    from templategate.core.model import ATTR_LAYOUT, Change

    change = Change(location="見積書!H:H", attribute=ATTR_LAYOUT,
                    old={"hidden": False}, new={"hidden": True},
                    detail="row or column changed: hidden False -> True")
    verdict, _ = classify(change, {})
    assert verdict == SUSPICIOUS


def test_an_edited_value_is_intended(quote):
    from templategate import diff

    baseline = snapshot(quote["baseline"])
    edited = [c for c in diff(quote["baseline"], quote["legit"])
              if c.location.endswith("D10")]
    assert edited
    verdict, _ = classify(edited[0], baseline)
    assert verdict == INTENDED


# --- selectors -----------------------------------------------------------

def test_contiguous_cells_collapse_into_one_range(quote):
    """A policy listing E10, E11, E12 separately is one nobody reads."""
    text = draft(quote["baseline"], quote["legit"])
    assert '"見積書!E10:E12"' in text
    assert '"見積書!E11"' not in text


def test_a_gap_in_the_run_starts_a_new_range(quote):
    """E13 is empty, so the cached results must not span across it."""
    text = draft(quote["baseline"], quote["legit"])
    assert '"見積書!E10:E16"' not in text


def test_a_whole_column_selector_survives_the_merger():
    """H:H is not a cell.  The merger has to carry it through untouched
    rather than raising on it, which is what openpyxl does when asked to
    read it as a coordinate."""
    from templategate.suggest import _merge_cells

    assert _merge_cells(["見積書!H:H"]) == ["見積書!H:H"]
    assert _merge_cells(["見積書!B2", "見積書!B3", "見積書!H:H"]) == [
        "見積書!B2:B3", "見積書!H:H"]


def test_one_reason_is_stated_once_for_the_whole_group(quote):
    text = draft(quote["baseline"], quote["legit"])
    assert text.count("a formula's cached result") == 1


# --- appending to a policy that already exists ---------------------------

def test_an_existing_policy_is_never_clobbered(quote, tmp_path):
    existing = tmp_path / "policy.yaml"
    original = (
        "version: 1\n"
        "target: excel\n"
        "mode: normal_input\n"
        "allow:\n"
        '  - selector: "見積書!B3"\n'
        "    attributes: [value]\n"
    )
    existing.write_text(original, encoding="utf-8")

    text = draft(quote["baseline"], quote["legit"], existing=existing)
    assert original.strip() in text


def test_additions_to_an_existing_policy_arrive_commented_out(quote, tmp_path):
    """Widening a live policy is a decision, so the draft cannot make it."""
    existing = tmp_path / "policy.yaml"
    existing.write_text(
        "version: 1\ntarget: excel\nmode: normal_input\n"
        'allow:\n  - selector: "見積書!B3"\n    attributes: [value]\n',
        encoding="utf-8")

    text = draft(quote["baseline"], quote["legit"], existing=existing)
    proposed = text.split("Proposed by", 1)[1]
    assert "D10" in proposed
    for line in proposed.splitlines():
        if "selector:" in line and "D10" in line:
            assert line.lstrip().startswith("#"), line


def test_an_existing_policy_still_loads_after_the_proposals(quote, tmp_path):
    existing = tmp_path / "policy.yaml"
    existing.write_text(
        "version: 1\ntarget: excel\nmode: normal_input\n"
        'allow:\n  - selector: "見積書!B3"\n    attributes: [value]\n',
        encoding="utf-8")

    merged = tmp_path / "merged.yaml"
    merged.write_text(draft(quote["baseline"], quote["legit"], existing=existing),
                      encoding="utf-8")

    loaded = load_policy(merged)
    assert [rule.selector for rule in loaded.allow] == ["見積書!B3"]


# --- nothing changed -----------------------------------------------------

def test_an_empty_diff_produces_a_draft_that_explains_itself(quote, tmp_path):
    text = draft(quote["baseline"], quote["baseline"])

    assert "nothing to learn" in text
    policy = tmp_path / "draft.yaml"
    policy.write_text(text, encoding="utf-8")
    loaded = load_policy(policy)
    assert loaded.target == "excel"


def test_a_draft_from_no_changes_denies_everything(quote, tmp_path):
    """An empty allow list is the strictest policy, not a broken one."""
    policy = tmp_path / "draft.yaml"
    policy.write_text(draft(quote["baseline"], quote["baseline"]),
                      encoding="utf-8")

    assert check(quote["baseline"], quote["baseline"], policy).passed
    assert not check(quote["baseline"], quote["legit"], policy).passed


# --- the command line ----------------------------------------------------

def test_the_draft_goes_to_stdout_by_default(quote, capsys):
    from templategate.cli import main

    assert main(["suggest", "--baseline", str(quote["baseline"]),
                 "--candidate", str(quote["legit"])]) == 0
    assert "allow:" in capsys.readouterr().out


def test_output_writes_a_file(quote, tmp_path, capsys):
    from templategate.cli import main

    destination = tmp_path / "draft.yaml"
    assert main(["suggest", "--baseline", str(quote["baseline"]),
                 "--candidate", str(quote["legit"]),
                 "--output", str(destination)]) == 0
    assert load_policy(destination).target == "excel"


def test_an_existing_output_file_is_not_overwritten(quote, tmp_path, capsys):
    """Same contract as init: a policy in the repo is not the tool's to lose."""
    from templategate.cli import main

    destination = tmp_path / "draft.yaml"
    destination.write_text("# hand written\n", encoding="utf-8")

    assert main(["suggest", "--baseline", str(quote["baseline"]),
                 "--candidate", str(quote["legit"]),
                 "--output", str(destination)]) == 2
    assert destination.read_text(encoding="utf-8") == "# hand written\n"
    assert "--force" in capsys.readouterr().err


def test_force_overwrites(quote, tmp_path, capsys):
    from templategate.cli import main

    destination = tmp_path / "draft.yaml"
    destination.write_text("# hand written\n", encoding="utf-8")

    assert main(["suggest", "--baseline", str(quote["baseline"]),
                 "--candidate", str(quote["legit"]),
                 "--output", str(destination), "--force"]) == 0
    assert "allow:" in destination.read_text(encoding="utf-8")


def test_writing_a_file_reminds_you_to_read_it(quote, tmp_path, capsys):
    from templategate.cli import main

    main(["suggest", "--baseline", str(quote["baseline"]),
          "--candidate", str(quote["legit"]),
          "--output", str(tmp_path / "draft.yaml")])
    assert "review it" in capsys.readouterr().err


# --- a document the gate could not fully read ---------------------------

def test_a_draft_from_a_damaged_document_says_so(quote, tmp_path):
    """Learning a policy from a file we could not parse teaches the wrong
    lesson, so the draft has to admit it up front."""
    from generate import rewrite_zip

    damaged = tmp_path / "damaged.xlsx"
    rewrite_zip(quote["legit"], damaged, drop=("xl/workbook.xml",))

    text = draft(quote["baseline"], damaged)
    assert "WARNING" in text
    assert "could not be read in full" in text


def test_an_unreadable_file_is_an_error_not_a_traceback(quote, tmp_path, capsys):
    from templategate.cli import main

    broken = tmp_path / "broken.xlsx"
    broken.write_bytes(b"not a zip at all")

    assert main(["suggest", "--baseline", str(quote["baseline"]),
                 "--candidate", str(broken)]) == 2
    assert "Traceback" not in capsys.readouterr().err

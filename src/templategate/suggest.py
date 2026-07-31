"""Draft a policy from an edit somebody has already looked at and approved.

Writing the first policy is the hard part, and it is hard for a reason that
has nothing to do with YAML: it takes knowing which of the twenty differences
between two spreadsheets are the edit, which are the collateral every save
produces, and which are the reason the gate exists.  That triage is expert
knowledge, and this module is where it is written down.

Three buckets, and the asymmetry between them is the whole design:

* **Intended** — the values and text the edit actually changed.  Allowed, with
  tight selectors and a comment saying what was observed.
* **Collateral** — differences that come from saving rather than editing:
  cached formula results, chart caches, a workbook default font, a column
  width read back.  Allowed, each with the explanation of why it is harmless.
* **Suspicious** — a formula replaced, an image gone, a sheet removed, a
  package part dropped.  **Never** allowed, whatever the observed edit did.
  They are listed at the top of the draft so a person decides consciously.

A draft that quietly allowed a destroyed formula because it happened to
appear in the sample edit would be worse than no draft at all.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from .api import diff_report, snapshot
from .core.model import (
    ATTR_CHARTS,
    ATTR_LAYOUT,
    ATTR_SHEET_SETTINGS,
    ATTR_TEXT,
    ATTR_VALUE,
    Change,
)
from .core.selector import _split_sheet, quote_sheet

# Attributes an edit is *made of*.  Everything else describes the document's
# shape rather than its content.
_CONTENT_ATTRIBUTES = frozenset({ATTR_VALUE, ATTR_TEXT})

INTENDED = "intended"
COLLATERAL = "collateral"
SUSPICIOUS = "suspicious"


def _cell_formula(snapshot_data: dict, location: str) -> object:
    """The formula a cell held in the baseline, if the location names one."""
    sheets = snapshot_data.get("sheets")
    if not sheets:
        return None
    sheet, separator, rest = _split_sheet(location)
    if separator != "!" or sheet not in sheets:
        return None
    return sheets[sheet]["cells"].get(rest, {}).get("formula")


def _changed_fields(change: Change) -> set[str]:
    """The named fields a delta-carrying change actually moved."""
    fields: set[str] = set()
    for side in (change.old, change.new):
        if isinstance(side, dict):
            fields |= set(side)
    return fields


def classify(change: Change, baseline: dict) -> tuple[str, str]:
    """Which bucket a change belongs in, and why — in words for the draft."""
    location, attribute = change.location, change.attribute

    if attribute == ATTR_VALUE and _cell_formula(baseline, location):
        return COLLATERAL, ("a formula's cached result; the formula itself is "
                            "unchanged, and whoever saved the file simply did "
                            "not store an answer for it")
    if attribute == ATTR_CHARTS:
        return COLLATERAL, "Excel refreshes chart caches on every save"
    if attribute == ATTR_SHEET_SETTINGS and _changed_fields(change) <= {
            "default_font.name", "default_font.size"}:
        return COLLATERAL, ("the workbook's default font, which each writing "
                            "tool records in its own way")
    if attribute == ATTR_LAYOUT and "hidden" not in _changed_fields(change):
        # Only the size moved.  Something being shown or hidden is a different
        # matter entirely, and stays suspicious — note that the *detail* says
        # "still hidden" in this case, so it cannot be read for this.
        return COLLATERAL, ("a row or column size read back by the editing "
                            "tool; nothing was shown or hidden")
    if attribute in _CONTENT_ATTRIBUTES:
        return INTENDED, "changed by the edit"
    return SUSPICIOUS, "not the kind of change an edit is made of"


# --- turning locations back into selectors -------------------------------

_SINGLE_CELL = re.compile(r"^\$?([A-Za-z]{1,3})\$?([1-9][0-9]{0,6})$")


def _cell_parts(location: str) -> tuple[str, int, int] | None:
    """Split ``Sheet!B7`` into its pieces, or give up.

    Anything that is not one plain cell — a whole column (``H:H``), a range,
    a named part — comes back as None and is carried through untouched.
    """
    sheet, separator, rest = _split_sheet(location)
    if separator != "!":
        return None
    match = _SINGLE_CELL.match(rest)
    if match is None:
        return None
    return sheet, column_index_from_string(match.group(1)), int(match.group(2))


def _merge_cells(locations: list[str]) -> list[str]:
    """Contiguous cells in one column become one range.

    A policy that lists C10, C11 and C12 separately is a policy nobody will
    read; C10:C12 is the same rule and says what it means.
    """
    columns: dict[tuple[str, int], list[int]] = {}
    leftovers: list[str] = []
    for location in locations:
        parts = _cell_parts(location)
        if parts is None:
            leftovers.append(location)
            continue
        sheet, column, row = parts
        columns.setdefault((sheet, column), []).append(row)

    selectors: list[str] = []
    for (sheet, column), rows in sorted(columns.items()):
        letter = get_column_letter(column)
        run_start = previous = None
        for row in sorted(rows):
            if run_start is None:
                run_start = previous = row
                continue
            if row == previous + 1:
                previous = row
                continue
            selectors.append(_range(sheet, letter, run_start, previous))
            run_start = previous = row
        if run_start is not None:
            selectors.append(_range(sheet, letter, run_start, previous))
    return selectors + sorted(set(leftovers))


def _range(sheet: str, letter: str, first: int, last: int) -> str:
    name = quote_sheet(sheet)
    if first == last:
        return f"{name}!{letter}{first}"
    return f"{name}!{letter}{first}:{letter}{last}"


def _merge_paragraphs(locations: list[str]) -> list[str]:
    """Consecutive body paragraphs become one p3-10 range."""
    numbers, leftovers = [], []
    for location in locations:
        if location.startswith("p") and location[1:].isdigit():
            numbers.append(int(location[1:]))
        else:
            leftovers.append(location)
    selectors: list[str] = []
    run_start = previous = None
    for number in sorted(numbers):
        if run_start is None:
            run_start = previous = number
            continue
        if number == previous + 1:
            previous = number
            continue
        selectors.append(f"p{run_start}" if run_start == previous
                         else f"p{run_start}-{previous}")
        run_start = previous = number
    if run_start is not None:
        selectors.append(f"p{run_start}" if run_start == previous
                         else f"p{run_start}-{previous}")
    return selectors + sorted(set(leftovers))


def _selectors_for(locations: list[str], target: str) -> list[str]:
    if target == "excel":
        return _merge_cells(locations)
    return _merge_paragraphs(locations)


def _rules(changes: list[Change], target: str) -> list[tuple[str, list[str], str]]:
    """(selector, attributes, comment) for one bucket of changes."""
    by_attribute: dict[str, list[str]] = {}
    reasons: dict[str, str] = {}
    for change, reason in changes:
        by_attribute.setdefault(change.attribute, []).append(change.location)
        reasons.setdefault(change.attribute, reason)

    rules = []
    for attribute, locations in sorted(by_attribute.items()):
        if attribute in _CONTENT_ATTRIBUTES or attribute == ATTR_LAYOUT:
            selectors = _selectors_for(sorted(set(locations)), target)
        else:
            # Package and workbook-level locations are already the selector.
            selectors = sorted(set(locations))
        for selector in selectors:
            rules.append((selector, [attribute], reasons[attribute]))
    return rules


# --- rendering the draft --------------------------------------------------

_HEADER = """\
# ==========================================================================
# TemplateGate policy draft
#
#   baseline : {baseline}
#   candidate: {candidate}
#
# This was generated from ONE observed edit.  It allows what that edit did
# and nothing else, so read every line before pinning it in CI — a draft is
# only ever as good as the example it was drawn from.
#
# The agent that edits the document must not be the one that runs
# `templategate suggest` to widen its own permissions.  Draft from an edit
# you have reviewed, then pin the result yourself.
# =========================================================================="""

_NOTHING_OBSERVED = """\
#
# The two documents are identical, so there was nothing to learn from them.
# What follows denies everything: a correct policy, and a useless one until
# you run suggest again against an edit that actually changed something."""


def _quote(text: str) -> str:
    return '"' + str(text).replace("\\", "\\\\").replace('"', '\\"') + '"'


def _wrap(text: str, width: int = 68) -> list[str]:
    words, lines, current = text.split(), [], ""
    for word in words:
        if current and len(current) + len(word) + 1 > width:
            lines.append(current)
            current = word
        else:
            current = f"{current} {word}".strip()
    if current:
        lines.append(current)
    return lines


def _rule_block(rules, indent: str = "  ") -> list[str]:
    """Render rules, stating each shared reason once above its group.

    Several ranges usually share one explanation — four separate copies of
    "this is a cached formula result" teaches the reader to skip comments.
    """
    lines: list[str] = []
    previous_reason = None
    for selector, attributes, reason in rules:
        if reason != previous_reason:
            if lines:
                lines.append("")
            for line in _wrap(f"observed: {reason}"):
                lines.append(f"{indent}# {line}")
            previous_reason = reason
        lines.append(f"{indent}- selector: {_quote(selector)}")
        lines.append(f"{indent}  attributes: [{', '.join(attributes)}]")
    return lines


def _warnings(buckets, degraded, changes) -> list[str]:
    lines: list[str] = []
    for role, reason in degraded.items():
        lines.append("#")
        for line in _wrap(f"WARNING: the {role} document could not be read in "
                          f"full ({reason}). A draft based on a document the "
                          "gate could not read is not trustworthy - fix the "
                          "file and run suggest again."):
            lines.append(f"# {line}")
    if not changes:
        lines.append(_NOTHING_OBSERVED)
    if buckets[SUSPICIOUS]:
        lines.append("#")
        lines.append("# NOT allowed by this draft - decide on each of these yourself:")
        for change, _reason in buckets[SUSPICIOUS]:
            what = change.detail or f"{change.attribute} changed"
            for line in _wrap(f"{change.location}: {what} - if this was "
                              "intended, add it consciously."):
                lines.append(f"#   {line}")
        lines.append("#")
        for line in _wrap("This draft therefore does not pass the edit it was "
                          "generated from. That is deliberate."):
            lines.append(f"# {line}")
    return lines


def _protect_block(buckets, target: str) -> list[str]:
    """Always-fail guards, minus anything this draft allows."""
    allowed = {change.attribute for change, _ in buckets[INTENDED]}
    allowed |= {change.attribute for change, _ in buckets[COLLATERAL]}
    guards = (["formula", "vba", "protection"] if target == "excel"
              else ["style", "paragraph_format", "content_control", "markup"])
    guards = [name for name in guards if name not in allowed]
    if not guards:
        return ["protect: []"]
    return [
        "# Redundant with default deny, but explicit - and a protect rule wins",
        "# even if an allow rule above ever grows to overlap it.",
        "protect:",
        '  - selector: "*"',
        f"    attributes: [{', '.join(guards)}]",
    ]


def _structural_block(target: str) -> list[str]:
    keys = (["sheets", "images", "defined_names", "charts", "pivot_tables",
             "drawings", "comments", "embedded", "custom_xml", "parts", "links"]
            if target == "excel" else
            ["images", "tables", "charts", "comments", "embedded",
             "custom_xml", "parts", "links"])
    return (["# Whole categories of content that must survive the edit.  Set one",
             "# to 'ignore' only once you have decided it may change.",
             "structural:"]
            + [f"  {key}: strict" for key in keys])


def draft(baseline, candidate, *, existing: Path | None = None) -> str:
    """A complete, commented policy drafted from one reviewed edit."""
    changes, degraded = diff_report(baseline, candidate)
    base_snapshot = snapshot(baseline)
    target = base_snapshot.get("target", "auto")

    buckets: dict[str, list] = {INTENDED: [], COLLATERAL: [], SUSPICIOUS: []}
    for change in changes:
        bucket, reason = classify(change, base_snapshot)
        buckets[bucket].append((change, reason))

    if existing is not None:
        return _append_to_existing(existing, buckets, target, baseline,
                                   candidate, degraded)

    lines = [_HEADER.format(baseline=baseline, candidate=candidate)]
    lines += _warnings(buckets, degraded, changes)
    lines += [
        "",
        "version: 1",
        f"target: {target}",
        "mode: normal_input",
        "",
        "# Anything not allowed below is a violation.  That is the point: the",
        "# list of allowed changes is the whole policy.",
        "allow:",
    ]
    allowed = _rules(buckets[INTENDED], target) + _rules(buckets[COLLATERAL], target)
    lines += _rule_block(allowed) if allowed else ["  []"]
    lines += [""] + _protect_block(buckets, target)
    lines += [""] + _structural_block(target)
    return "\n".join(lines) + "\n"


def _append_to_existing(existing: Path, buckets, target, baseline, candidate,
                        degraded) -> str:
    """The policy as it stands, plus what the observed edit would still need.

    Additions arrive commented out: adopting one has to be a decision, not a
    side effect of running a command.
    """
    from .core.evaluator import evaluate
    from .core.policy import load_policy

    policy = load_policy(existing)
    uncovered = []
    for bucket in (INTENDED, COLLATERAL):
        changes = [change for change, _ in buckets[bucket]]
        _allowed, violations = evaluate(changes, policy)
        blocked = {violation.change.location for violation in violations}
        uncovered += [(change, reason) for change, reason in buckets[bucket]
                      if change.location in blocked]

    lines = [existing.read_text(encoding="utf-8").rstrip(), ""]
    lines.append("# " + "=" * 72)
    lines.append("# Proposed by `templategate suggest` from an observed edit:")
    lines.append(f"#   baseline : {baseline}")
    lines.append(f"#   candidate: {candidate}")
    lines += _warnings(buckets, degraded, uncovered or buckets[SUSPICIOUS])
    if not uncovered:
        lines += ["#",
                  "# Nothing to add: this policy already allows every change",
                  "# the edit made that a draft would be willing to allow.",
                  "# " + "=" * 72]
        return "\n".join(lines) + "\n"
    lines += ["#",
              "# Uncomment the ones you agree with, under `allow:` above.",
              "# " + "=" * 72]
    lines += ["# " + line for line in _rule_block(_rules(uncovered, target), "")]
    return "\n".join(lines) + "\n"

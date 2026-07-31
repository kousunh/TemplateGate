"""Excel (.xlsx / .xlsm) snapshot extraction.

A snapshot is a plain dict of everything TemplateGate compares:
values, formulas, formats, merges, conditional formatting, data validation,
sheet structure, images, headers/footers, print settings, defined names and
the VBA project hash.  Extraction is read-only.
"""

from __future__ import annotations

import datetime
import hashlib
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from openpyxl.worksheet.worksheet import Worksheet

from ..core.package import list_part_names, take_package_snapshot

_DEFAULT_CELL = {"value": None, "formula": None, "format": None}


def _scalar(value):
    """A colour component only if it is a real value.

    openpyxl leaves unset descriptors on the object, and stringifying one
    produces noise like "Integer[('name', 'theme')]" that then shows up in
    every colour diff.  Anything that is not a plain value is simply not set.
    """
    if isinstance(value, bool) or isinstance(value, (str, int, float)):
        return value
    return None


def _rgb(value) -> str | None:
    """An ARGB string with its alpha canonicalised.

    openpyxl writes opaque colours as 00RRGGBB and Excel writes the same
    colour as FFRRGGBB.  Both mean fully opaque, so comparing the raw string
    reports a colour change every time a library-authored file is opened and
    saved in Excel.  Only the two opaque spellings are folded together; a
    genuinely translucent colour keeps its alpha.
    """
    text = _scalar(value)
    if text is None:
        return None
    text = str(text)
    if len(text) == 8 and text[:2].upper() in ("00", "FF"):
        return "FF" + text[2:].upper()
    return text


def _color(c) -> str | None:
    """A colour as one readable token, so a diff reads FF000000 -> FFFFFFFF."""
    if c is None:
        return None
    rgb = _rgb(getattr(c, "rgb", None))
    if rgb is not None:
        return rgb
    theme = _scalar(getattr(c, "theme", None))
    if theme is not None:
        tint = _scalar(getattr(c, "tint", None)) or 0
        return f"theme{theme}" if not tint else f"theme{theme}/tint{tint}"
    indexed = _scalar(getattr(c, "indexed", None))
    if indexed is not None:
        return f"indexed{indexed}"
    return None


def _side(s) -> tuple | None:
    if s is None or s.style is None:
        return None
    return (s.style, _color(s.color))


def _default_font(workbook) -> tuple:
    """The font a cell gets when it does not ask for one.

    Excel materialises this onto every cell on save, and it is
    locale-dependent — a Japanese Excel writes ＭＳ Ｐゴシック where a library
    wrote nothing at all.  Resolving each cell against its own workbook's
    default makes the comparison say the same thing on any machine.
    """
    fonts = getattr(workbook, "_fonts", None)
    if not fonts:
        return (None, None)
    first = fonts[0]
    return (_plain(getattr(first, "name", None)), _plain(getattr(first, "sz", None)))


def _format_key(cell, default_font: tuple = (None, None)) -> tuple | None:
    """A cell's style as named fields.  None == no style at all.

    Named rather than positional so a difference can be reported as
    "numfmt '#,##0' -> '#,##0,'" instead of an unreadable tuple.  Kept as a
    tuple of pairs so it stays hashable and stable in JSON.
    """
    if not cell.has_style:
        return None
    f, fill, b, a = cell.font, cell.fill, cell.border, cell.alignment
    fields: dict[str, object] = {}

    def record(name, value, *, default=None) -> None:
        # Only what this cell actually says.  A style carries a value for
        # every property whether or not it means anything, and listing all of
        # them turns "this cell went bold" into twenty rows of noise.
        if value is not None and value != default:
            fields[name] = value

    # A cell that merely inherits the workbook default is not styled, however
    # explicitly the file happens to spell that out.
    record("font.name", _plain(f.name), default=default_font[0])
    record("font.size", _plain(f.size), default=default_font[1])
    record("font.bold", bool(f.bold), default=False)
    record("font.italic", bool(f.italic), default=False)
    record("font.underline", _plain(f.underline))
    record("font.strike", bool(f.strike), default=False)
    record("font.color", _color(f.color))

    pattern = _plain(fill.patternType)
    record("fill.pattern", pattern)
    if pattern:  # a fill with no pattern paints nothing, whatever its colours
        record("fill.foreground", _color(getattr(fill, "fgColor", None)))
        record("fill.background", _color(getattr(fill, "bgColor", None)))

    for edge in ("left", "right", "top", "bottom", "diagonal"):
        record(f"border.{edge}", _side(getattr(b, edge, None)))

    record("align.horizontal", _plain(a.horizontal))
    record("align.vertical", _plain(a.vertical))
    record("align.wrap_text", bool(a.wrap_text), default=False)
    record("align.rotation", _plain(a.text_rotation), default=0)
    record("align.indent", _plain(a.indent), default=0)
    record("numfmt", _plain(cell.number_format), default="General")
    # Locked is the default and only bites once the sheet is protected, which
    # is compared separately; an *unlocked* cell is the notable one.
    record("protect.locked", _plain(cell.protection.locked), default=True)
    record("protect.hidden", _plain(cell.protection.hidden), default=False)
    # A style that sets nothing beyond the workbook defaults is not a style.
    # Excel materialises the default font onto every cell on save, which makes
    # `has_style` true for cells that still say exactly nothing.
    return tuple(sorted(fields.items())) if fields else None


def _plain(value):
    """Reduce a cell value to something comparable and JSON-safe.

    An openpyxl object must never reach the snapshot.  ArrayFormula and
    DataTableFormula define no __eq__, so two loads of the same untouched file
    compare unequal and every round-trip reports a phantom change; worse, they
    are not str, so a formula wearing one would be filed under the `value`
    attribute and slip past a policy that protects `formula`.
    """
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    if isinstance(value, datetime.timedelta):
        return value.total_seconds()
    # Anything unforeseen: describe it by its public attributes, never by
    # repr(), which would bake in a memory address.
    fields = getattr(value, "__dict__", None)
    if fields:
        return f"{type(value).__name__}" + repr(
            sorted((k, str(v)) for k, v in fields.items() if not k.startswith("_"))
        )
    return f"{type(value).__name__}:{value}"


def _rich_runs(value) -> tuple | None:
    """Per-run formatting of a rich-text cell, or None for a plain one.

    A cell can hold several differently formatted runs; the cell-level style
    describes none of them, so turning one run white leaves every other
    fingerprint untouched.
    """
    if not isinstance(value, CellRichText):
        return None
    runs = []
    for item in value:
        if isinstance(item, TextBlock):
            f = item.font
            runs.append((
                item.text, _plain(f.b), _plain(f.i), _plain(f.u), _plain(f.sz),
                _plain(f.rFont), _rgb(getattr(f.color, "rgb", None)
                                      if f.color is not None else None),
            ))
        else:
            runs.append((str(item), None, None, None, None, None, None))
    return tuple(runs)


def _formula_of(value):
    """The canonical formula carried by a cell value, or None."""
    if isinstance(value, ArrayFormula):
        return ("array", value.ref, value.text)
    if isinstance(value, DataTableFormula):
        return ("data_table",) + tuple(sorted(dict(value).items()))
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def _populated(ws):
    """The cells a sheet actually has, in a stable order.

    ``iter_rows`` walks the sheet's *declared* dimension, so a single value
    parked in XFD100000 makes it materialize a billion empty cells and the
    gate appears to hang.  openpyxl already keeps only the cells that exist,
    keyed by (row, column); reading that directly makes the cost track the
    content instead of the geometry.
    """
    grid = getattr(ws, "_cells", None)
    if grid is None:  # pragma: no cover - every Worksheet has one
        return [cell for row in ws.iter_rows() for cell in row]
    return [grid[key] for key in sorted(grid)]


def _cells(ws_formula, ws_value, default_font: tuple = (None, None)) -> dict:
    cells: dict[str, dict] = {}
    value_grid = getattr(ws_value, "_cells", {})
    for cell in _populated(ws_formula):
        formula = _formula_of(cell.value)
        runs = _rich_runs(cell.value)
        fmt = _format_key(cell, default_font)
        if runs is not None:
            fmt = (fmt or ()) + (("runs", runs),)
        if formula:
            cached_cell = value_grid.get((cell.row, cell.column))
            cached = cached_cell.value if cached_cell is not None else None
        else:
            cached = cell.value
        cached = _plain(cached)
        if cached is None and formula is None and fmt is None:
            continue
        cells[cell.coordinate] = {
            "value": cached,
            "formula": formula,
            "format": fmt,
        }
    return cells


def _dxf_key(dxf) -> tuple | None:
    """The styling a conditional-format rule applies when it fires.

    Without this a rule keeps its condition and loses its highlight — the
    warning still 'works', it just no longer shows.
    """
    if dxf is None:
        return None
    font, fill, border = dxf.font, dxf.fill, dxf.border
    return (
        ("font", getattr(font, "b", None), getattr(font, "i", None),
         _color(getattr(font, "color", None))) if font is not None else None,
        ("fill", getattr(fill, "patternType", None),
         _color(getattr(fill, "fgColor", None)),
         _color(getattr(fill, "bgColor", None))) if fill is not None else None,
        ("border", _side(getattr(border, "left", None)),
         _side(getattr(border, "right", None)),
         _side(getattr(border, "top", None)),
         _side(getattr(border, "bottom", None))) if border is not None else None,
        ("numfmt", getattr(dxf.numFmt, "formatCode", None)
         if dxf.numFmt is not None else None),
    )


def _conditional_formatting(ws) -> dict:
    out: dict[str, list] = {}
    for cf in ws.conditional_formatting:
        rules = []
        for rule in cf.rules:
            rules.append((
                rule.type,
                getattr(rule, "operator", None),
                tuple(getattr(rule, "formula", None) or []),
                getattr(rule, "text", None),
                bool(getattr(rule, "stopIfTrue", False)),
                _dxf_key(getattr(rule, "dxf", None)),
            ))
        out[str(cf.sqref)] = sorted(map(repr, rules))
    return out


def _or_default(value, default):
    """An attribute Excel omitted still means what the schema says it means.

    Excel drops attributes that hold their default value, so a rule written
    by a library with errorStyle="stop" comes back with no errorStyle at all.
    Filling the documented default makes the two spellings equal — while a
    real downgrade to "information" is an explicit value and stays visible.
    """
    return default if value is None else value


def _data_validation(ws) -> dict:
    out: dict[str, tuple] = {}
    for dv in ws.data_validations.dataValidation:
        out[str(dv.sqref)] = (
            dv.type, _or_default(dv.operator, "between"),
            dv.formula1, dv.formula2,
            bool(dv.allowBlank), bool(dv.showDropDown),
            # How hard the rule pushes back.  Turning errorStyle from "stop"
            # to "information" leaves the rule in place but lets anything through.
            _or_default(dv.errorStyle, "stop"),
            bool(dv.showErrorMessage), bool(dv.showInputMessage),
            dv.errorTitle, dv.error, dv.promptTitle, dv.prompt,
        )
    return out


def _protection(ws) -> tuple:
    p = ws.protection
    return tuple(
        _plain(getattr(p, name, None))
        for name in ("sheet", "objects", "scenarios", "formatCells",
                     "formatColumns", "formatRows", "insertColumns",
                     "insertRows", "deleteColumns", "deleteRows",
                     "selectLockedCells", "selectUnlockedCells", "sort",
                     "autoFilter", "pivotTables")
    ) + (bool(getattr(p, "password", None)),)


def _sheet_settings(ws) -> tuple:
    """Sheet-level view state that changes what a reader can see or do."""
    return (
        ("freeze_panes", _plain(ws.freeze_panes)),
        ("auto_filter", _plain(ws.auto_filter.ref)),
        ("tab_color", _color(ws.sheet_properties.tabColor)),
        ("gridlines", _plain(getattr(ws.sheet_view, "showGridLines", None))),
        ("zoom", _plain(getattr(ws.sheet_view, "zoomScale", None))),
        ("right_to_left", _plain(getattr(ws.sheet_view, "rightToLeft", None))),
    )


def _layout(ws) -> dict:
    """Row heights and column widths, and which of them are hidden.

    A hidden row is not a deleted row: the numbers are still in the file and
    still feed every formula, they just stop being on the page.
    """
    out: dict[str, tuple] = {}
    for index, dim in ws.row_dimensions.items():
        if dim.hidden or dim.height is not None:
            out[f"{index}:{index}"] = (bool(dim.hidden), _plain(dim.height))
    for letter, dim in ws.column_dimensions.items():
        if dim.hidden or dim.width is not None:
            out[f"{letter}:{letter}"] = (bool(dim.hidden), _plain(dim.width))
    return out


def _images(ws) -> list[dict]:
    images = []
    for img in getattr(ws, "_images", []):
        try:
            data = img._data()
        except Exception:
            data = b""
        anchor = getattr(img.anchor, "_from", None)
        images.append({
            "sha256": hashlib.sha256(data).hexdigest(),
            "anchor": (anchor.col, anchor.row) if anchor is not None else None,
            "size": (round(img.width or 0), round(img.height or 0)),
        })
    return sorted(images, key=lambda d: (d["sha256"], d["anchor"] or (-1, -1)))


def _header_footer(ws) -> dict:
    out = {}
    for name in ("oddHeader", "oddFooter", "evenHeader", "evenFooter",
                 "firstHeader", "firstFooter"):
        hf = getattr(ws, name, None)
        if hf is None:
            continue
        parts = {side: getattr(hf, side).text for side in ("left", "center", "right")
                 if getattr(hf, side).text}
        if parts:
            out[name] = parts
    return out


def _print_settings(ws) -> dict:
    ps, pm = ws.page_setup, ws.page_margins
    return {
        "orientation": ps.orientation,
        "paper_size": ps.paperSize,
        "scale": ps.scale,
        "fit_to_width": ps.fitToWidth,
        "fit_to_height": ps.fitToHeight,
        "print_area": str(ws.print_area) if ws.print_area else None,
        "print_title_rows": ws.print_title_rows,
        "print_title_cols": ws.print_title_cols,
        "margins": (pm.left, pm.right, pm.top, pm.bottom, pm.header, pm.footer),
    }


_EMPTY_SHEET = {
    "cells": {},
    "merges": [],
    "conditional_formatting": {},
    "data_validation": {},
    "images": [],
    "header_footer": {},
    "print": {},
    "protection": (),
    "settings": (),
    "layout": {},
    "defined_names": {},
}


def _sheet_snapshot(ws, index: int, wb_value, *,
                    ignored_errors: str | None = None,
                    default_font: tuple = (None, None)) -> dict:
    """One sheet's contents.

    A workbook may also contain chartsheets, which carry no cell grid at all —
    asking one for its rows raises AttributeError, so they are recorded by
    presence and kind only.  A worksheet swapped for a chartsheet still shows
    up, as a change of kind.
    """
    common = {
        "index": index,
        "visibility": ws.sheet_state,
        "kind": "worksheet" if isinstance(ws, Worksheet) else "chartsheet",
    }
    if not isinstance(ws, Worksheet):
        return {**common, **_EMPTY_SHEET}
    return {
        **common,
        "cells": _cells(ws, wb_value[ws.title], default_font),
        "merges": sorted(str(m) for m in ws.merged_cells.ranges),
        "conditional_formatting": _conditional_formatting(ws),
        "data_validation": _data_validation(ws),
        "images": _images(ws),
        "header_footer": _header_footer(ws),
        "print": _print_settings(ws),
        "protection": _protection(ws),
        "settings": _sheet_settings(ws) + (("ignored_errors", ignored_errors),),
        "layout": _layout(ws),
        # Sheet-scoped names are a separate namespace from the workbook's:
        # two sheets may each define "Rates" pointing somewhere different.
        "defined_names": {name: dn.value
                          for name, dn in ws.defined_names.items()},
    }


_MAIN_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_DOC_REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
_PKG_REL_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"


def _ignored_errors(path: Path) -> dict[str, str]:
    """Per sheet, the cell errors Excel has been told not to flag.

    Suppressing the warning triangles is how a broken formula or a number
    stored as text stops looking broken, and openpyxl does not model the
    element at all — so it is read from the sheet part directly.
    """
    import xml.etree.ElementTree as ElementTree
    import zipfile

    try:
        with zipfile.ZipFile(path) as zf:
            names = set(zf.namelist())
            if "xl/workbook.xml" not in names:
                return {}
            workbook = ElementTree.fromstring(zf.read("xl/workbook.xml"))
            targets = {}
            if "xl/_rels/workbook.xml.rels" in names:
                rels = ElementTree.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
                for relationship in rels.iter(_PKG_REL_NS + "Relationship"):
                    targets[relationship.get("Id")] = relationship.get("Target", "")

            found: dict[str, str] = {}
            for sheet in workbook.iter(_MAIN_NS + "sheet"):
                target = targets.get(sheet.get(_DOC_REL_NS + "id"), "")
                if not target:
                    continue
                # Targets come both absolute ("/xl/worksheets/sheet1.xml")
                # and relative to the workbook part ("worksheets/sheet1.xml").
                part = target.lstrip("/")
                if not part.startswith("xl/"):
                    part = "xl/" + part
                if part not in names:
                    continue
                root = ElementTree.fromstring(zf.read(part))
                blocks = [ElementTree.tostring(element, encoding="unicode")
                          for element in root.iter(_MAIN_NS + "ignoredErrors")]
                if blocks:
                    found[sheet.get("name", "")] = hashlib.sha256(
                        "".join(blocks).encode("utf-8")).hexdigest()[:12]
            return found
    except Exception:
        # Never let an optional extra stop the rest of the snapshot.
        return {}


def take_snapshot(path: str | Path) -> dict:
    path = Path(path)
    wb_formula = load_workbook(path, data_only=False, rich_text=True)
    wb_value = load_workbook(path, data_only=True)

    suppressed = _ignored_errors(path)
    default_font = _default_font(wb_formula)
    sheets: dict[str, dict] = {}
    for index, name in enumerate(wb_formula.sheetnames):
        ws = wb_formula[name]
        sheets[name] = _sheet_snapshot(ws, index, wb_value,
                                       ignored_errors=suppressed.get(name),
                                       default_font=default_font)

    defined_names = {}
    for name, dn in wb_formula.defined_names.items():
        defined_names[name] = dn.value

    return {
        "target": "excel",
        "format": path.suffix.lstrip(".").lower(),
        "sheets": sheets,
        "defined_names": defined_names,
        "settings": (("calc_mode", _plain(wb_formula.calculation.calcMode)),
                     ("full_calc_on_load",
                      _plain(wb_formula.calculation.fullCalcOnLoad))),
        "package": take_package_snapshot(path),
        "part_names": list_part_names(path),
    }
